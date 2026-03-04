"""Tracker data processor for student progress tracking.

Transforms raw CSV form responses into a comprehensive Excel workbook
with multiple tabs for different priority levels and a summary dashboard.
"""

import csv
import io
import re
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from services.file_processor import FileProcessor, ProcessingResult

import discord  # For embed creation


# ==================== Data Classes ====================

@dataclass
class SubmissionsResult:
    """Result of a submissions check operation."""
    success: bool
    summary_embed: Optional[discord.Embed] = None
    error_message: Optional[str] = None
    total_enrolled: int = 0
    submitted_count: int = 0
    missing_count: int = 0
    at_risk_students: List[Dict] = field(default_factory=list)
    flagged_students: List[Dict] = field(default_factory=list)
    on_track_students: List[Dict] = field(default_factory=list)

@dataclass
class StudentRecord:
    """Represents a processed student record with all calculated fields."""
    # Core identifiers
    student_id: str = ""
    name: str = ""
    member_id: str = ""
    discord_username: str = ""
    email: str = ""
    phone: str = ""
    
    # Time tracking
    week: int = 0
    submission_num: int = 0  # Sequential submission number (Wed W1=1, Sun W1=2, Wed W2=3, etc.)
    submission_date: str = ""
    wed_submitted: bool = False
    sun_submitted: bool = False
    submission_count_cumulative: int = 0
    
    # Phase tracking
    current_phase: str = ""
    weeks_in_phase: int = 1
    contribution_num: int = 1
    contribution_start_week: int = 1
    weeks_on_contribution: int = 1
    weeks_remaining: int = 10
    timeline_type: str = "Standard"
    phase_changed_this_week: bool = False
    
    # Links
    readme_link: str = ""
    issue_url: str = ""
    fork_url: str = ""
    mr_url: str = ""
    
    # Deliverables completion
    why_chosen_complete: bool = False
    reproduction_complete: bool = False
    solution_complete: bool = False
    implementation_complete: bool = False
    testing_complete: bool = False
    feedback_complete: bool = False
    deliverables_expected: int = 0
    deliverables_complete: int = 0
    
    # Git activity
    commits_this_week: int = 0
    last_commit_date: str = ""
    days_since_commit: int = 0
    total_commits: int = 0
    
    # MR tracking
    mr_status: str = ""
    mr_created_date: str = ""
    comment_count: int = 0
    has_maintainer_feedback: bool = False
    
    # Progress
    progress_summary: str = ""
    next_week_plan: str = ""
    blocked: bool = False
    blocker_desc: str = ""
    support_requested: str = ""
    
    # Issue tracking
    issue_url_previous_week: str = ""
    issue_changed: bool = False
    issue_change_week: int = 0
    issue_swap_detected: bool = False
    new_contribution_detected: bool = False
    
    # Grading and intervention
    grade_status: str = "🟢 ON TRACK"
    intervention_type: str = ""
    intervention_sent_date: str = ""
    consecutive_misses: int = 0
    last_week_wed_missing: bool = False  # Missing Wednesday from previous week
    last_week_sun_missing: bool = False  # Missing Sunday from previous week
    member_id_mismatch: bool = False  # Member ID column doesn't match "What's your Member ID?" column
    invalid_member_id: bool = False  # Member ID is empty, #N/A, or otherwise invalid
    
    # Office hours
    tue_office_hours: bool = False
    thu_office_hours: bool = False
    wed_lecture: bool = False
    
    # Notes
    cam_notes: str = ""
    
    # Raw data for reference
    raw_data: Dict[str, Any] = field(default_factory=dict)


# ==================== Column Mappings ====================

# Maps CSV column headers to StudentRecord fields
# Note: "Week" is computed from start_date, not read from typeform
CSV_COLUMN_MAP = {
    "#": "student_id",
    "What's your name?": "name",
    "What's your Member ID?": "member_id",
    "Member ID": "member_id",
    "What is your Discord username?": "discord_username",
    "Which week is this?": "week",
    "Which contribution are you reporting on?": "contribution_num",
    "Link to your contribution README": "readme_link",
    "Which submission are you completing?": "_submission_type",
    "Submission for": "_submission_day",
    "What phase are you currently in?": "current_phase",
    "What phase are you currently in?_2": "current_phase",  # Duplicate column
    "Direct link to your GitLab issue": "issue_url",
    "Have you completed the \"Why I chose this issue\" section in your README?": "why_chosen_complete",
    "Direct link to your GitLab fork": "fork_url",
    "Have you documented your reproduction process in your README?": "reproduction_complete",
    "Have you documented your solution approach in your README?": "solution_complete",
    "Have you documented your implementation progress in your README?": "implementation_complete",
    "Have you documented your testing strategy in your README?": "testing_complete",
    "Direct link to your Merge Request (MR) or Pull Request (PR)": "mr_url",
    "Have you documented any maintainer feedback in your README?": "feedback_complete",
    "Briefly summarize what you accomplished this week": "progress_summary",
    "What's your plan for next week?": "next_week_plan",
    "Are you currently blocked or stuck?": "blocked",
    "Describe what you're blocked on": "blocker_desc",
    "What kind of support would help you most right now?": "support_requested",
    "Submitted At": "submission_date",
    "Date Submitted": "submission_date",
    "Tags": "_tags",
}

# Alternative column names for discord username (in case of different form versions)
DISCORD_USERNAME_COLUMNS = [
    "What is your Discord username?",
    "Discord Username",
    "Discord",
    "discord_username",
]


# Master CSV column mappings
MASTER_CSV_COLUMNS = {
    "member_id": ["Member ID", "member_id", "MemberID"],
    "discord_username": ["Discord Username", "Discord", "discord_username"],
    "full_name": ["Full Name", "Name", "full_name"],
    "email": ["Email", "email", "Secondary Email"],
    "phone": ["Phone", "phone", "Phone Number", "phone_number", "Mobile", "Cell"],
    "slack_username": ["Slack Username", "Slack", "slack_username"],
    "status": ["Status", "status"],
    "university": ["University", "university"],
    "github": ["GitLab Username", "Github", "GitHub", "github", "gitlab_username"],
    "cohort": ["Cohort Name", "Cohort", "cohort"],
    "location": ["Location", "Section", "location"],
}


def _normalize_header(header: str) -> str:
    """Normalize a header for flexible matching."""
    return header.strip().lower().replace("?", "").rstrip()


def _get_value_flexible(row: Dict[str, str], target_col: str) -> Optional[str]:
    """Get value from row using flexible header matching.
    
    Tries exact match first, then falls back to normalized matching.
    """
    # Try exact match first
    if target_col in row:
        return row[target_col]
    
    # Try normalized matching
    target_normalized = _normalize_header(target_col)
    for row_col, value in row.items():
        if _normalize_header(row_col) == target_normalized:
            return value
    
    return None


def _preprocess_master_csv(master_text: str) -> str:
    """Preprocess master CSV to find actual header row and strip metadata.
    
    The master CSV may have metadata rows at the top (dates, week info, etc.)
    before the actual header row. This function finds the header row
    (containing "Member ID") and returns the CSV starting from that row.
    
    Also handles leading empty columns by stripping them.
    
    Args:
        master_text: Raw CSV text
        
    Returns:
        Cleaned CSV text starting from the header row
    """
    lines = master_text.splitlines()
    header_row_idx = None
    
    # Find the row containing "Member ID" (the actual header)
    for idx, line in enumerate(lines):
        if "Member ID" in line or "member_id" in line.lower():
            header_row_idx = idx
            break
    
    if header_row_idx is None:
        # No header found, return original text
        return master_text
    
    # Get lines from header onwards
    data_lines = lines[header_row_idx:]
    
    # Check if first column is empty (line starts with comma)
    # and strip it from all rows
    if data_lines and data_lines[0].startswith(','):
        cleaned_lines = []
        for line in data_lines:
            if line.startswith(','):
                line = line[1:]  # Remove leading comma
            cleaned_lines.append(line)
        data_lines = cleaned_lines
    
    return '\n'.join(data_lines)


def _preprocess_typeform_csv(typeform_text: str) -> str:
    """Preprocess typeform CSV to find actual header row and strip metadata.
    
    The typeform CSV may have metadata rows at the top before the actual header.
    This function finds the header row (containing key columns like "Member ID",
    "Week", "Submitted At") and returns the CSV starting from that row.
    
    Also handles duplicate column names by making them unique (appending _2, _3, etc.)
    
    Args:
        typeform_text: Raw CSV text
        
    Returns:
        Cleaned CSV text starting from the header row
    """
    lines = typeform_text.splitlines()
    header_row_idx = None
    
    # Find the row containing typeform header columns
    # Look for rows with "Member ID" AND other typeform-specific columns
    for idx, line in enumerate(lines):
        line_lower = line.lower()
        # Check for typeform-specific header indicators
        has_member_id = "member id" in line_lower or "member_id" in line_lower
        has_week = "week" in line_lower
        has_submitted = "submitted" in line_lower or "date" in line_lower
        
        # Typeform header should have multiple of these
        if has_member_id and (has_week or has_submitted):
            header_row_idx = idx
            break
    
    if header_row_idx is None:
        # No header found, return original text
        return typeform_text
    
    # Get lines from header onwards
    data_lines = lines[header_row_idx:]
    
    # Check if first column is empty (line starts with comma) and strip it
    # But be careful - typeform may have legitimate empty first cells
    # Only strip if the header row itself starts with comma
    if data_lines and data_lines[0].startswith(','):
        cleaned_lines = []
        for line in data_lines:
            if line.startswith(','):
                line = line[1:]  # Remove leading comma
            cleaned_lines.append(line)
        data_lines = cleaned_lines
    
    # Handle duplicate column names in header row
    # Use csv module to properly parse header (handles escaped quotes like "")
    if data_lines:
        header_line = data_lines[0]
        
        # Parse header using csv module for proper quote handling
        import csv as csv_module
        reader = csv_module.reader([header_line])
        header_parts = next(reader)
        
        # Make duplicate column names unique
        seen_cols = {}
        unique_parts = []
        for col in header_parts:
            col_stripped = col.strip()
            if col_stripped in seen_cols:
                seen_cols[col_stripped] += 1
                # Create unique name by appending suffix
                new_col = f"{col_stripped}_{seen_cols[col_stripped]}"
                unique_parts.append(new_col)
            else:
                seen_cols[col_stripped] = 1
                unique_parts.append(col_stripped)
        
        # Reconstruct header line - quote fields that contain commas or quotes
        def quote_field(field):
            if ',' in field or '"' in field:
                return '"' + field.replace('"', '""') + '"'
            return field
        
        data_lines[0] = ','.join(quote_field(f) for f in unique_parts)
    
    return '\n'.join(data_lines)


# ==================== Style Definitions ====================

class Styles:
    """Excel style definitions."""
    
    # Fills
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    RED_FILL = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    ORANGE_FILL = PatternFill(start_color="FFB347", end_color="FFB347", fill_type="solid")
    YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    LIGHT_YELLOW_FILL = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
    GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    LIGHT_GREEN_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    DASHBOARD_HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    DASHBOARD_SECTION_FILL = PatternFill(start_color="D6DCE5", end_color="D6DCE5", fill_type="solid")
    
    # Fonts
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    BOLD_FONT = Font(bold=True)
    TITLE_FONT = Font(bold=True, size=14)
    DASHBOARD_TITLE_FONT = Font(bold=True, size=16, color="FFFFFF")
    
    # Alignment
    CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
    LEFT_ALIGN = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # Border
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )


# ==================== Tracker Processor ====================

class TrackerDataProcessor(FileProcessor):
    """Processes tracker CSV data into a comprehensive Excel workbook.
    
    Creates multiple tabs:
    - Tab 1: Intervention Tracker (all fields)
    - Tab 2: At Risk Students (P1 Priority)
    - Tab 3: Flagged Students (P2 Priority)  
    - Tab 4: On Track Students (P3 Spot Checks)
    - Tab 5: Weekly Summary Dashboard
    """
    
    # Output column definitions for each tab
    ALL_COLUMNS = [
        "student_id", "name", "member_id", "discord_username", "email", "phone", "week",
        "submission_date", "wed_submitted", "sun_submitted", "submission_count_cumulative",
        "current_phase", "weeks_in_phase", "contribution_num", "contribution_start_week",
        "weeks_on_contribution", "weeks_remaining", "timeline_type", "phase_changed_this_week",
        "readme_link", "issue_url", "fork_url", "mr_url",
        "why_chosen_complete", "reproduction_complete", "solution_complete",
        "implementation_complete", "testing_complete", "feedback_complete",
        "deliverables_expected", "deliverables_complete",
        "commits_this_week", "last_commit_date", "days_since_commit", "total_commits",
        "mr_status", "mr_created_date", "comment_count", "has_maintainer_feedback",
        "progress_summary", "next_week_plan", "blocked", "blocker_desc", "support_requested",
        "issue_url_previous_week", "issue_changed", "issue_change_week",
        "issue_swap_detected", "new_contribution_detected",
        "grade_status", "intervention_type", "intervention_sent_date", "consecutive_misses",
        "tue_office_hours", "thu_office_hours", "wed_lecture", "cam_notes"
    ]
    
    AT_RISK_COLUMNS = [
        "name", "week", "current_phase", "weeks_in_phase", "timeline_type",
        "sun_submitted", "consecutive_misses", "deliverables_complete",
        "commits_this_week", "blocked", "intervention_type", "readme_link", "cam_notes"
    ]
    
    FLAGGED_COLUMNS = [
        "name", "week", "current_phase", "weeks_in_phase", "timeline_type",
        "deliverables_complete", "commits_this_week", "days_since_commit",
        "blocked", "intervention_type", "readme_link"
    ]
    
    ON_TRACK_COLUMNS = [
        "name", "week", "current_phase", "weeks_in_phase", "submission_count_cumulative",
        "mr_status", "progress_summary", "cam_notes"
    ]
    
    @property
    def input_type(self) -> str:
        return "csv"
    
    @property
    def output_type(self) -> str:
        return "xlsx"
    
    def _get_week_deadlines(self, start_date: datetime, week_num: int) -> Tuple[datetime, datetime]:
        """Calculate Wednesday and Sunday deadlines for a given week.
        
        Args:
            start_date: Program start date (assumed to be a Monday)
            week_num: Week number (1-based)
            
        Returns:
            Tuple of (wednesday_deadline, sunday_deadline) as datetime objects
        """
        # Week N starts at: start_date + (N-1) * 7 days
        week_start = start_date + timedelta(days=(week_num - 1) * 7)
        
        # Calculate days to Wednesday (weekday 2) and Sunday (weekday 6)
        start_weekday = start_date.weekday()  # Monday=0, Sunday=6
        days_to_wed = (2 - start_weekday) % 7
        days_to_sun = (6 - start_weekday) % 7
        if days_to_sun == 0:
            days_to_sun = 7  # If start is Sunday, next Sunday is 7 days away
        
        wed_deadline = week_start + timedelta(days=days_to_wed)
        sun_deadline = week_start + timedelta(days=days_to_sun)
        
        return wed_deadline, sun_deadline
    
    def _map_early_submission_week(
        self, 
        submission_date: Optional[datetime], 
        typeform_week: int,
        is_sunday: bool,
        start_date: datetime,
        target_date: datetime,
        current_week: int
    ) -> Tuple[int, bool]:
        """Map an early submission to its effective week and check visibility.
        
        Early submissions (before start_date) are mapped to Week 1.
        Visibility is based on whether the deadline has passed for that submission type.
        
        Args:
            submission_date: When the typeform was actually submitted
            typeform_week: Week value from typeform input
            is_sunday: Whether this is a Sunday submission (vs Wednesday)
            start_date: Program start date
            target_date: Date the report is being run
            current_week: Computed current week based on target_date
            
        Returns:
            Tuple of (effective_week, is_visible)
            - effective_week: The week this submission should be counted for
            - is_visible: Whether this submission should be included in the report
        """
        # Determine effective week
        effective_week = typeform_week if typeform_week > 0 else current_week
        
        is_early_submission = submission_date and submission_date.date() < start_date.date()
        
        if is_early_submission:
            # Early submission - map to Week 1
            effective_week = 1
        
        # Check visibility based on deadline
        # EXCEPTION: Early submissions (before start_date) are ALWAYS visible
        # since the student submitted proactively before the program started
        if is_early_submission:
            is_visible = True
        else:
            wed_deadline, sun_deadline = self._get_week_deadlines(start_date, effective_week)
            
            if is_sunday:
                # Sunday submissions visible after Sunday deadline
                is_visible = target_date.date() >= sun_deadline.date()
            else:
                # Wednesday submissions visible after Wednesday deadline
                is_visible = target_date.date() >= wed_deadline.date()
        
        return effective_week, is_visible
    
    def process(self, data: bytes, options: Optional[Dict[str, Any]] = None) -> ProcessingResult:
        """Process CSV data into multi-tab Excel workbook.
        
        Args:
            data: Typeform CSV data bytes
            options: Optional dict with:
                - master_data: Master roster CSV bytes (optional)
                - zoom_data: Zoom attendance CSV bytes (optional)
        """
        options = options or {}
        
        try:
            # Parse typeform CSV
            text_data = data.decode('utf-8-sig')
            text_data = _preprocess_typeform_csv(text_data)
            
            # Auto-detect delimiter (handles both CSV and TSV)
            sample = text_data[:4096]  # Sample first 4KB for sniffing
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=',\t;|')
            except csv.Error:
                # Default to comma if sniffing fails
                dialect = 'excel'
            
            csv_reader = csv.DictReader(io.StringIO(text_data), dialect=dialect)
            raw_rows = list(csv_reader)
            
            # Filter by date and apply early submission visibility logic
            if options.get('filter_by_date') and options.get('target_date'):
                target_date = options['target_date']
                start_date = options.get('start_date')
                current_week = options.get('current_week', 1)
                filtered_rows = []
                
                # Log deadline info for current week
                if start_date:
                    wed_dl, sun_dl = self._get_week_deadlines(start_date, current_week)
                    print(f"[TrackerProcessor] Week {current_week} deadlines: Wed {wed_dl.strftime('%m/%d/%Y')}, Sun {sun_dl.strftime('%m/%d/%Y')}")
                    print(f"[TrackerProcessor] Running report as of: {target_date.strftime('%m/%d/%Y')}")
                
                # Find submission date column and submission type column
                if raw_rows:
                    headers = list(raw_rows[0].keys())
                    date_col = None
                    for col in ["Submitted At", "Date Submitted", "submission_date", "Submit Date", "Submit Date (UTC)"]:
                        if col in headers:
                            date_col = col
                            break
                    
                    # Find submission type column
                    submission_type_col = None
                    for col in ["Which submission are you completing?", "Submission for"]:
                        if col in headers:
                            submission_type_col = col
                            break
                    
                    # Find week column
                    week_col = None
                    for col in ["Which week is this?", "Week"]:
                        if col in headers:
                            week_col = col
                            break
                    
                    if date_col:
                        early_mapped = 0
                        hidden_sun = 0
                        hidden_wed = 0
                        
                        for row in raw_rows:
                            date_str = row.get(date_col, "")
                            submission_dt = None
                            if date_str:
                                for fmt in ["%m/%d/%Y %H:%M:%S", "%m/%d/%Y %H:%M", "%Y-%m-%d %H:%M:%S", 
                                           "%Y-%m-%dT%H:%M:%S", "%m/%d/%Y", "%Y-%m-%d",
                                           "%d-%b-%y", "%d-%b-%Y"]:
                                    try:
                                        submission_dt = datetime.strptime(date_str.strip(), fmt)
                                        break
                                    except ValueError:
                                        continue
                            
                            # Determine if this is a Sunday submission
                            is_sunday = False
                            if submission_type_col:
                                sub_type = str(row.get(submission_type_col, "")).lower()
                                is_sunday = "sunday" in sub_type or sub_type == "sun"
                            
                            # Get typeform week
                            typeform_week = 0
                            if week_col:
                                try:
                                    week_str = str(row.get(week_col, "")).replace("Week ", "").strip()
                                    typeform_week = int(week_str)
                                except:
                                    pass
                            
                            # Apply early submission mapping and visibility check
                            if start_date and submission_dt:
                                effective_week, is_visible = self._map_early_submission_week(
                                    submission_dt, typeform_week, is_sunday,
                                    start_date, target_date, current_week
                                )
                                
                                # Track early submissions mapped to Week 1
                                if submission_dt.date() < start_date.date():
                                    early_mapped += 1
                                
                                if is_visible:
                                    # Store effective week for later use
                                    row['_effective_week'] = effective_week
                                    filtered_rows.append(row)
                                else:
                                    # Track why it was hidden
                                    if is_sunday:
                                        hidden_sun += 1
                                    else:
                                        hidden_wed += 1
                            elif submission_dt is None or submission_dt.date() <= target_date.date():
                                # Fallback: include if on or before target date
                                filtered_rows.append(row)
                        
                        raw_rows = filtered_rows
                        print(f"[TrackerProcessor] Filtered to {len(raw_rows)} rows (visible as of {target_date.strftime('%m/%d/%Y')})")
                        if early_mapped > 0:
                            print(f"[TrackerProcessor] {early_mapped} early submissions mapped to Week 1")
                        if hidden_sun > 0 or hidden_wed > 0:
                            print(f"[TrackerProcessor] Hidden (deadline not passed): {hidden_wed} Wed, {hidden_sun} Sun")
            
            # Build discord username lookup from master CSV (primary source)
            discord_lookup: Dict[str, str] = {}
            master_data = options.get('master_data')
            if master_data:
                discord_lookup = self._build_master_discord_lookup(master_data)
            
            # Build phone lookup from app CSV
            phone_lookup: Dict[str, str] = {}
            app_data = options.get('app_data')
            if app_data:
                phone_lookup = self._build_app_phone_lookup(app_data)
            
            # If no rows after filtering, but we have master data, continue with just At Risk students
            if not raw_rows:
                if options.get('filter_by_date') and master_data:
                    # No typeform submissions before target date - all enrolled students are At Risk
                    students = []
                    current_week = options.get('current_week', 1)
                    start_date = options.get('start_date')
                    students = self._add_missing_students_as_at_risk(
                        students, master_data, discord_lookup, options['target_date'], 
                        current_week, start_date, phone_lookup
                    )
                    
                    if not students:
                        return ProcessingResult(
                            success=False,
                            error_message="No students found in master CSV"
                        )
                else:
                    return ProcessingResult(
                        success=False,
                        error_message="CSV file is empty"
                    )
            else:
                # Supplement with typeform discord data (fills gaps if master doesn't have entry)
                typeform_discord_lookup = self._build_discord_lookup(raw_rows)
                for member_id, discord_name in typeform_discord_lookup.items():
                    if member_id not in discord_lookup:
                        discord_lookup[member_id] = discord_name
                
                # Build name lookup for fallback matching when Member ID is invalid
                name_lookup = {}
                if master_data:
                    name_lookup = self._build_name_lookup_from_master(master_data)
                
                # Build contact lookup for email and phone
                contact_lookup = {}
                if master_data:
                    contact_lookup = self._build_master_contact_lookup(master_data)
                
                # Merge phone numbers from app CSV (overrides master if present)
                app_data = options.get('app_data')
                if app_data:
                    phone_lookup = self._build_app_phone_lookup(app_data)
                    for member_id, phone in phone_lookup.items():
                        if member_id in contact_lookup:
                            contact_lookup[member_id]['phone'] = phone
                        else:
                            contact_lookup[member_id] = {'discord': '', 'email': '', 'phone': phone}
                
                # Transform to StudentRecord objects
                students = self._transform_records(raw_rows, discord_lookup, name_lookup, contact_lookup)
                
                # Apply effective week from early submission mapping
                # The _effective_week was calculated during filtering based on:
                # - Early submissions (before start_date) → mapped to Week 1
                # - Visibility filtering based on Wed/Sun deadlines
                current_week = options.get('current_week', 1)
                for i, student in enumerate(students):
                    if i < len(raw_rows) and '_effective_week' in raw_rows[i]:
                        # Use the pre-calculated effective week from visibility filtering
                        student.week = raw_rows[i]['_effective_week']
                    elif student.week and student.week != current_week:
                        # User indicated a different week in typeform
                        # Keep their indicated week
                        pass
                    else:
                        # Use computed week
                        student.week = current_week
                
                # Get phase completions early so we can use them in derived fields
                phase_completions = options.get('phase_completions', {})
                
                # Calculate derived fields (pass phase_completions for MISSING_PREVIOUS_PHASE check)
                self._calculate_derived_fields(students, phase_completions)
                
                # Apply manual phase completions (overrides typeform phase if higher)
                if phase_completions:
                    self._apply_phase_completions(students, phase_completions)
                
                # Determine grade status and interventions
                # Pass start_date and target_date so we can check per-student deadlines
                start_date = options.get('start_date')
                target_date = options.get('target_date')
                bypasses = options.get('bypasses', {})
                
                self._calculate_grade_status(students, start_date=start_date, target_date=target_date, bypasses=bypasses)
                
                # Mark typeform-only students (not in master CSV) as MISSING_ADMISSION_INFO
                if master_data:
                    self._mark_typeform_only_students(students, master_data)
                
                # If filtering by date, add students with NO submissions as "At Risk"
                if options.get('filter_by_date') and options.get('target_date') and master_data:
                    current_week = options.get('current_week', 1)
                    start_date = options.get('start_date')
                    students = self._add_missing_students_as_at_risk(
                        students, master_data, discord_lookup, options['target_date'], 
                        current_week, start_date, phone_lookup
                    )
                
                # Enrich with GitLab data if service is provided
                gitlab_service = options.get('gitlab_service')
                if gitlab_service:
                    target_date = options.get('target_date')
                    validate_commits = options.get('validate_commits', False)
                    validate_mrs = options.get('validate_mrs', False)
                    nofilter = options.get('nofilter', False)
                    
                    # Build github username lookup from master CSV
                    github_lookup = {}
                    if master_data:
                        github_lookup = self._build_github_lookup(master_data)
                    
                    self._enrich_with_gitlab(
                        students, gitlab_service, target_date,
                        validate_commits=validate_commits,
                        validate_mrs=validate_mrs,
                        nofilter=nofilter,
                        github_lookup=github_lookup
                    )
            
            # Create workbook
            wb = Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Create tabs (Master first, then priority tabs)
            start_date = options.get('start_date')
            target_date = options.get('target_date')
            self._create_master_tab(wb, students)
            self._create_at_risk_tab(wb, students, start_date, target_date)
            self._create_flagged_tab(wb, students, start_date, target_date)
            self._create_on_track_tab(wb, students, start_date, target_date)
            self._create_summary_tab(wb, students)
            
            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            # Count unique students by member_id (or name as fallback)
            unique_students = len(set(s.member_id or s.name for s in students))
            
            return ProcessingResult(
                success=True,
                output_data=output.read(),
                output_filename="tracker_report.xlsx",
                rows_processed=unique_students,
                students=students  # Include raw student records for autogroup
            )
            
        except Exception as e:
            return ProcessingResult(
                success=False,
                error_message=f"Processing error: {e}"
            )
    
    def process_submissions(self, typeform_data: bytes, master_data: bytes,
                           start_date: datetime, target_date: datetime,
                           current_week: int, app_data: Optional[bytes] = None) -> SubmissionsResult:
        """Process submissions for real-time checking.
        
        Args:
            typeform_data: Typeform CSV data bytes
            master_data: Master roster CSV data bytes
            start_date: Program start date
            target_date: Date to filter submissions up to
            current_week: Calculated current week number
            app_data: Optional app CSV data bytes for phone numbers
            
        Returns:
            SubmissionsResult with summary embed and student lists
        """
        try:
            # Parse master CSV to get all enrolled students
            master_text = master_data.decode('utf-8-sig')
            master_text = _preprocess_master_csv(master_text)
            sample = master_text[:4096]
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=',\t;|')
            except csv.Error:
                dialect = 'excel'
            
            master_reader = csv.DictReader(io.StringIO(master_text), dialect=dialect)
            master_rows = list(master_reader)
            
            if not master_rows:
                return SubmissionsResult(
                    success=False,
                    error_message="Master CSV is empty"
                )
            
            # Build enrolled students lookup
            headers = list(master_rows[0].keys())
            member_id_col = self._find_column(headers, MASTER_CSV_COLUMNS["member_id"])
            name_col = self._find_column(headers, MASTER_CSV_COLUMNS["full_name"])
            discord_col = self._find_column(headers, MASTER_CSV_COLUMNS["discord_username"])
            email_col = self._find_column(headers, MASTER_CSV_COLUMNS["email"])
            phone_col = self._find_column(headers, MASTER_CSV_COLUMNS["phone"])
            
            enrolled_students: Dict[str, Dict] = {}
            for row in master_rows:
                member_id = str(row.get(member_id_col, "")).strip() if member_id_col else ""
                if member_id:
                    enrolled_students[member_id] = {
                        'member_id': member_id,
                        'name': str(row.get(name_col, "")).strip() if name_col else "",
                        'discord': str(row.get(discord_col, "")).strip() if discord_col else "",
                        'email': str(row.get(email_col, "")).strip() if email_col else "",
                        'phone': str(row.get(phone_col, "")).strip() if phone_col else "",
                        'submissions': [],
                        'issues': []
                    }
            
            # Merge phone numbers from app CSV (overrides master if present)
            if app_data:
                phone_lookup = self._build_app_phone_lookup(app_data)
                for member_id, phone in phone_lookup.items():
                    if member_id in enrolled_students:
                        enrolled_students[member_id]['phone'] = phone
            
            # Parse typeform CSV and filter by date
            typeform_text = typeform_data.decode('utf-8-sig')
            typeform_text = _preprocess_typeform_csv(typeform_text)
            sample = typeform_text[:4096]
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=',\t;|')
            except csv.Error:
                dialect = 'excel'
            
            typeform_reader = csv.DictReader(io.StringIO(typeform_text), dialect=dialect)
            typeform_rows = list(typeform_reader)
            
            # Get submission date column
            submission_date_col = None
            if typeform_rows:
                tf_headers = list(typeform_rows[0].keys())
                for col in ["Submitted At", "Date Submitted", "submission_date", "Submit Date", "Submit Date (UTC)"]:
                    if col in tf_headers:
                        submission_date_col = col
                        break
            
            # Find submission type column
            submission_type_col = None
            if typeform_rows:
                tf_headers = list(typeform_rows[0].keys())
                for col in ["Which submission are you completing?", "Submission for"]:
                    if col in tf_headers:
                        submission_type_col = col
                        break
            
            # Process typeform submissions with early submission mapping
            submitted_member_ids = set()
            invalid_member_id_entries = []  # Track entries with invalid Member IDs
            
            # Build name lookup for fallback matching (name -> member_id)
            name_to_member_id = {}
            for mid, student in enrolled_students.items():
                student_name = student.get('name', '').strip().lower()
                if student_name:
                    name_to_member_id[student_name] = mid
            
            for row in typeform_rows:
                # Parse submission date
                date_str = row.get(submission_date_col, "") if submission_date_col else ""
                submission_dt = None
                if date_str:
                    for fmt in ["%m/%d/%Y %H:%M:%S", "%m/%d/%Y %H:%M", "%Y-%m-%d %H:%M:%S", 
                               "%Y-%m-%dT%H:%M:%S", "%m/%d/%Y", "%Y-%m-%d",
                               "%d-%b-%y", "%d-%b-%Y"]:
                        try:
                            submission_dt = datetime.strptime(date_str.strip(), fmt)
                            break
                        except ValueError:
                            continue
                
                # Get member ID - check "Member ID" column first
                member_id_col_value = str(row.get("Member ID", "") or row.get("member_id", "")).strip()
                whats_member_id_value = str(row.get("What's your Member ID?", "")).strip()
                
                # Check if values are invalid
                invalid_values = ['#N/A', 'N/A', 'NULL', 'NONE', '#REF!', '#VALUE!', '-', '']
                primary_is_invalid = not member_id_col_value or member_id_col_value.upper() in invalid_values
                secondary_is_invalid = not whats_member_id_value or whats_member_id_value.upper() in invalid_values
                
                # Determine which member_id to use and if there's an invalid ID issue
                member_id = None
                has_invalid_member_id = False
                matched_by_name = False
                
                if not primary_is_invalid and member_id_col_value in enrolled_students:
                    # Primary column is valid and matches enrolled student
                    member_id = member_id_col_value
                elif not secondary_is_invalid and whats_member_id_value in enrolled_students:
                    # Try fallback to "What's your Member ID?" column
                    member_id = whats_member_id_value
                    # Mark as having invalid Member ID since primary column was bad
                    if primary_is_invalid:
                        has_invalid_member_id = True
                else:
                    # Both Member ID columns failed - try name matching as last resort
                    # Try multiple possible name column variations
                    name = None
                    for name_col in ["What's your name?", "Name", "name", "Full Name", "full_name", "Student Name"]:
                        name = row.get(name_col, "")
                        if name and str(name).strip():
                            break
                    
                    name_clean = str(name).strip().lower() if name else ""
                    
                    if name_clean:
                        # Try to find matching enrolled student by name
                        if name_clean in name_to_member_id:
                            member_id = name_to_member_id[name_clean]
                            has_invalid_member_id = True
                            matched_by_name = True
                
                if not member_id:
                    # Can't match to any enrolled student - track as invalid entry
                    name = row.get("What's your name?", "") or row.get("Name", "")
                    invalid_member_id_entries.append({
                        'name': str(name).strip() if name else "Unknown",
                        'member_id_raw': member_id_col_value or "empty",
                        'row': row
                    })
                    continue
                
                # Mark the student if they have invalid Member ID issue
                if has_invalid_member_id:
                    enrolled_students[member_id]['invalid_member_id'] = True
                    if matched_by_name:
                        enrolled_students[member_id]['invalid_member_id_value'] = f"{member_id_col_value or 'empty'} (matched by name)"
                    else:
                        enrolled_students[member_id]['invalid_member_id_value'] = member_id_col_value or "empty"
                
                # Determine submission type (Wed/Sun)
                is_sunday = False
                if submission_type_col:
                    sub_type = str(row.get(submission_type_col, "")).lower()
                    is_sunday = "sunday" in sub_type or sub_type == "sun"
                
                # Get week number from submission
                week_str = row.get("Which week is this?", "") or row.get("Week", "")
                typeform_week = 0
                if week_str:
                    try:
                        typeform_week = int(str(week_str).replace("Week ", "").strip())
                    except:
                        pass
                
                # Apply early submission mapping and visibility check
                effective_week, is_visible = self._map_early_submission_week(
                    submission_dt, typeform_week, is_sunday,
                    start_date, target_date, current_week
                )
                
                if not is_visible:
                    continue
                
                submitted_member_ids.add(member_id)
                
                # Check for Member ID mismatch
                member_id_col_value = row.get("Member ID", "")
                whats_member_id_value = row.get("What's your Member ID?", "")
                if member_id_col_value and whats_member_id_value:
                    if str(member_id_col_value).strip() != str(whats_member_id_value).strip():
                        enrolled_students[member_id]['member_id_mismatch'] = True
                
                # Track submission info with effective week
                enrolled_students[member_id]['submissions'].append({
                    'week': effective_week,
                    'date': submission_dt,
                    'phase': row.get("What phase are you currently in?", ""),
                    'wed': not is_sunday,
                    'sun': is_sunday
                })
            
            # Calculate deadline dates for display (current week)
            wed_deadline, sun_deadline = self._get_week_deadlines(start_date, current_week)
            wed_deadline_passed = target_date.date() >= wed_deadline.date()
            sun_deadline_passed = target_date.date() >= sun_deadline.date()
            
            # Categorize students
            # Sunday is the official weekly submission (required)
            # Wednesday is a mid-week check-in (optional)
            at_risk = []
            flagged = []
            on_track = []
            missing_students = []
            
            for member_id, student in enrolled_students.items():
                subs = student['submissions']
                issues = []
                
                if not subs:
                    # No submissions at all - AT RISK if any deadline has passed
                    week1_wed, week1_sun = self._get_week_deadlines(start_date, 1)
                    if target_date.date() >= week1_wed.date():
                        # Wednesday or Sunday deadline passed - AT RISK
                        issues.append("No submissions")
                        at_risk.append({**student, 'issues': issues, 'status': 'AT RISK'})
                    else:
                        # No deadlines passed yet
                        on_track.append({**student, 'issues': [], 'status': 'ON TRACK'})
                    continue
                
                # Find the weeks this student has submissions for
                weeks_with_subs = set(s['week'] for s in subs if s['week'] > 0)
                
                # Check each week's Sunday submissions (Wednesday is optional if they have any submission)
                missing_sunday = False
                
                for week in weeks_with_subs:
                    week_wed_dl, week_sun_dl = self._get_week_deadlines(start_date, week)
                    week_subs = [s for s in subs if s['week'] == week]
                    has_sun = any(s['sun'] for s in week_subs)
                    
                    # Check if Sunday deadline passed but no Sunday submission
                    # Only flag missing Sunday - Wednesday is optional if they have any submission
                    if target_date.date() >= week_sun_dl.date() and not has_sun:
                        issues.append(f"Missing Sunday submission (Week {week})")
                        missing_sunday = True
                
                # Check for member ID mismatch
                has_member_id_mismatch = student.get('member_id_mismatch', False)
                if has_member_id_mismatch:
                    issues.append("Member ID mismatch")
                
                # Check for invalid Member ID in submission (primary column was #N/A, empty, etc.)
                has_invalid_member_id = student.get('invalid_member_id', False)
                if has_invalid_member_id:
                    invalid_val = student.get('invalid_member_id_value', 'invalid')
                    issues.append(f"Invalid Member ID column: {invalid_val}")
                
                # Categorize based on issues
                # AT RISK: missing Sunday or invalid member ID (critical issues)
                # FLAGGED: member ID mismatch only (data quality concern)
                if missing_sunday or has_invalid_member_id:
                    at_risk.append({**student, 'issues': issues, 'status': 'AT RISK'})
                elif has_member_id_mismatch:
                    flagged.append({**student, 'issues': issues, 'status': 'FLAGGED'})
                else:
                    on_track.append({**student, 'issues': [], 'status': 'ON TRACK'})
            
            # Add entries with invalid Member IDs to at_risk
            for entry in invalid_member_id_entries:
                at_risk.append({
                    'member_id': f"INVALID: {entry['member_id_raw']}",
                    'name': entry['name'],
                    'discord': '',
                    'submissions': [],
                    'issues': [f"Invalid Member ID: {entry['member_id_raw']}"],
                    'status': 'AT RISK'
                })
            
            # Create summary embed
            embed = discord.Embed(
                title="📊 Submission Status Report",
                description=f"**Week {current_week}** (as of {target_date.strftime('%m/%d/%Y')})",
                color=discord.Color.blue()
            )
            
            # Show deadline status
            wed_status = "✅ Passed" if wed_deadline_passed else "⏳ Not yet"
            sun_status = "✅ Passed" if sun_deadline_passed else "⏳ Not yet"
            
            embed.add_field(
                name="📅 Deadlines",
                value=(
                    f"**Wednesday ({wed_deadline.strftime('%m/%d')}):** {wed_status}\n"
                    f"**Sunday ({sun_deadline.strftime('%m/%d')}):** {sun_status}"
                ),
                inline=False
            )
            
            total = len(enrolled_students)
            submitted = len(on_track)
            
            embed.add_field(
                name="📈 Overview",
                value=(
                    f"**Total Enrolled:** {total}\n"
                    f"**On Track:** {len(on_track)} ({len(on_track)*100//total if total else 0}%)\n"
                    f"**Flagged:** {len(flagged)}\n"
                    f"**At Risk:** {len(at_risk)}"
                ),
                inline=False
            )
            
            # List at-risk students (max 10)
            # Sort to show more important issues first, "No submissions" last
            if at_risk:
                def at_risk_sort_key(s):
                    issues_str = ', '.join(s['issues'])
                    # "No submissions" goes last (high number)
                    if "No submissions" in issues_str:
                        return 2
                    # Other issues go first (low number)
                    return 1
                
                sorted_at_risk = sorted(at_risk, key=at_risk_sort_key)
                at_risk_list = "\n".join([
                    f"• {s['name']} ({s['member_id']}): {', '.join(s['issues'])}"
                    for s in sorted_at_risk[:10]
                ])
                if len(at_risk) > 10:
                    at_risk_list += f"\n... and {len(at_risk) - 10} more"
                embed.add_field(
                    name="🔴 At Risk",
                    value=at_risk_list or "None",
                    inline=False
                )
            
            # List flagged students (max 10)
            if flagged:
                flagged_list = "\n".join([
                    f"• {s['name']} ({s['member_id']}): {', '.join(s['issues'])}"
                    for s in flagged[:10]
                ])
                if len(flagged) > 10:
                    flagged_list += f"\n... and {len(flagged) - 10} more"
                embed.add_field(
                    name="🟡 Flagged",
                    value=flagged_list or "None",
                    inline=False
                )
            
            embed.set_footer(text=f"Use !tracker submissions_download for full report")
            
            return SubmissionsResult(
                success=True,
                summary_embed=embed,
                total_enrolled=total,
                submitted_count=len(on_track),
                missing_count=len(at_risk) + len(flagged),
                at_risk_students=at_risk,
                flagged_students=flagged,
                on_track_students=on_track
            )
            
        except Exception as e:
            return SubmissionsResult(
                success=False,
                error_message=f"Error processing submissions: {e}"
            )
    
    def _find_column(self, headers: List[str], possible_names: List[str]) -> Optional[str]:
        """Find a column by checking possible names (case-sensitive first, then insensitive)."""
        # Exact match first
        for name in possible_names:
            if name in headers:
                return name
        
        # Case-insensitive fallback
        headers_lower = {h.lower(): h for h in headers}
        for name in possible_names:
            if name.lower() in headers_lower:
                return headers_lower[name.lower()]
        
        return None
    
    def _add_missing_students_as_at_risk(
        self, 
        students: List[StudentRecord], 
        master_data: bytes, 
        discord_lookup: Dict[str, str],
        target_date: datetime,
        current_week: int = 1,
        start_date: Optional[datetime] = None,
        phone_lookup: Optional[Dict[str, str]] = None
    ) -> List[StudentRecord]:
        """Add students from master CSV with no submissions as At Risk entries.
        
        Args:
            students: Existing list of student records from typeform
            master_data: Master roster CSV bytes
            discord_lookup: Member ID to discord username mapping
            target_date: Target date for filtering
            current_week: Current week number (based on start_date and target_date)
            start_date: Program start date (for calculating missed deadlines)
            phone_lookup: Optional member_id -> phone mapping from app CSV
            
        Returns:
            Updated list including At Risk entries for students with no submissions
        """
        phone_lookup = phone_lookup or {}
        try:
            # Get all member_ids that already have submissions
            submitted_member_ids = set(s.member_id for s in students if s.member_id)
            
            # Calculate consecutive misses (all Wed/Sun deadlines missed before target_date)
            consecutive_misses = 0
            if start_date:
                consecutive_misses = self._calculate_missed_deadlines(start_date, target_date)
            
            # Parse master CSV
            master_text = master_data.decode('utf-8-sig')
            master_text = _preprocess_master_csv(master_text)
            sample = master_text[:4096]
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=',\t;|')
            except csv.Error:
                dialect = 'excel'
            
            master_reader = csv.DictReader(io.StringIO(master_text), dialect=dialect)
            master_rows = list(master_reader)
            
            if not master_rows:
                return students
            
            # Find column names
            headers = list(master_rows[0].keys())
            member_id_col = self._find_column(headers, MASTER_CSV_COLUMNS["member_id"])
            name_col = self._find_column(headers, MASTER_CSV_COLUMNS["full_name"])
            discord_col = self._find_column(headers, MASTER_CSV_COLUMNS["discord_username"])
            email_col = self._find_column(headers, MASTER_CSV_COLUMNS["email"])
            phone_col = self._find_column(headers, MASTER_CSV_COLUMNS["phone"])
            
            # Add missing students as At Risk
            for row in master_rows:
                member_id = str(row.get(member_id_col, "")).strip() if member_id_col else ""
                
                if member_id and member_id not in submitted_member_ids:
                    # This student has no submissions - add as At Risk
                    name = str(row.get(name_col, "")).strip() if name_col else ""
                    discord = str(row.get(discord_col, "")).strip() if discord_col else ""
                    email = str(row.get(email_col, "")).strip() if email_col else ""
                    phone = str(row.get(phone_col, "")).strip() if phone_col else ""
                    
                    # Use discord lookup as fallback
                    if not discord and member_id in discord_lookup:
                        discord = discord_lookup[member_id]
                    
                    # Use phone lookup from app CSV (overrides master)
                    if member_id in phone_lookup:
                        phone = phone_lookup[member_id]
                    
                    # Create an At Risk record for this student
                    # Use current_week (based on start_date and target_date) instead of 0
                    at_risk_record = StudentRecord(
                        member_id=member_id,
                        name=name,
                        discord_username=discord,
                        email=email,
                        phone=phone,
                        week=current_week,
                        submission_date="",
                        wed_submitted=False,
                        sun_submitted=False,
                        submission_count_cumulative=0,
                        current_phase="",
                        weeks_in_phase=0,
                        contribution_num=0,
                        contribution_start_week=0,
                        weeks_on_contribution=0,
                        weeks_remaining=0,
                        timeline_type="Critical",
                        grade_status="🔴 AT RISK",
                        intervention_type="NO_SUBMISSIONS",
                        consecutive_misses=consecutive_misses
                    )
                    
                    students.append(at_risk_record)
            
            return students
            
        except Exception as e:
            print(f"[TrackerProcessor] Error adding missing students: {e}")
            return students
    
    def _mark_typeform_only_students(
        self,
        students: List[StudentRecord],
        master_data: bytes
    ) -> None:
        """Mark students who submitted typeform but are NOT in the master CSV as At Risk.
        
        These students have submitted progress updates but their member_id doesn't
        appear in the master roster, indicating missing admission information.
        
        Args:
            students: List of student records from typeform
            master_data: Master roster CSV bytes
        """
        try:
            # Parse master CSV to get all enrolled member_ids
            master_text = master_data.decode('utf-8-sig')
            master_text = _preprocess_master_csv(master_text)
            sample = master_text[:4096]
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=',\t;|')
            except csv.Error:
                dialect = 'excel'
            
            master_reader = csv.DictReader(io.StringIO(master_text), dialect=dialect)
            master_rows = list(master_reader)
            
            if not master_rows:
                return
            
            # Find member_id column
            headers = list(master_rows[0].keys())
            member_id_col = self._find_column(headers, MASTER_CSV_COLUMNS["member_id"])
            
            if not member_id_col:
                print("[TrackerProcessor] Master CSV: Member ID column not found for typeform-only check")
                return
            
            # Build set of all member_ids in master CSV
            master_member_ids = set()
            for row in master_rows:
                member_id = str(row.get(member_id_col, "")).strip()
                if member_id:
                    master_member_ids.add(member_id)
            
            # Check each typeform student and mark those not in master as MISSING_ADMISSION_INFO
            marked_count = 0
            for student in students:
                member_id = str(student.member_id).strip()
                if member_id and member_id not in master_member_ids:
                    student.grade_status = "🔴 AT RISK"
                    student.intervention_type = "MISSING_ADMISSION_INFO"
                    marked_count += 1
            
            if marked_count > 0:
                print(f"[TrackerProcessor] Marked {marked_count} typeform-only students as MISSING_ADMISSION_INFO")
            
        except Exception as e:
            print(f"[TrackerProcessor] Error checking typeform-only students: {e}")
    
    def _build_github_lookup(self, master_data: bytes) -> Dict[str, str]:
        """Build a member_id -> github_username lookup from master roster CSV.
        
        Args:
            master_data: Raw bytes of master CSV file
            
        Returns:
            Dict mapping member_id to github_username
        """
        github_lookup: Dict[str, str] = {}
        
        try:
            text_data = master_data.decode('utf-8-sig')
            text_data = _preprocess_master_csv(text_data)
            sample = text_data[:4096]
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=',\t;|')
            except csv.Error:
                dialect = 'excel'
            
            csv_reader = csv.DictReader(io.StringIO(text_data), dialect=dialect)
            rows = list(csv_reader)
            
            if not rows:
                return github_lookup
            
            headers = list(rows[0].keys())
            
            # Find member_id column
            member_id_col = self._find_column(headers, MASTER_CSV_COLUMNS["member_id"])
            if not member_id_col:
                print("[TrackerProcessor] Master CSV: Member ID column not found for GitHub lookup")
                return github_lookup
            
            # Find github username column
            github_col = self._find_column(headers, MASTER_CSV_COLUMNS["github"])
            if not github_col:
                print("[TrackerProcessor] Master CSV: Github column not found")
                return github_lookup
            
            # Build lookup
            for row in rows:
                member_id = str(row.get(member_id_col, "")).strip()
                github_username = str(row.get(github_col, "")).strip()
                
                if member_id and github_username:
                    github_lookup[member_id] = github_username.lower()
            
            print(f"[TrackerProcessor] Master CSV: Built GitHub lookup with {len(github_lookup)} entries")
            
        except Exception as e:
            print(f"[TrackerProcessor] Error parsing master CSV for GitHub lookup: {e}")
        
        return github_lookup
    
    def _enrich_with_gitlab(
        self,
        students: List[StudentRecord],
        gitlab_service,
        target_date: Optional[datetime],
        validate_commits: bool = False,
        validate_mrs: bool = False,
        nofilter: bool = False,
        github_lookup: Optional[Dict[str, str]] = None
    ) -> None:
        """Enrich student records with GitLab API data.
        
        Modes:
        - nofilter: Just checks existence of commits/MRs, no ownership validation
        - validate_commits: Flags AT_RISK if commits are on repos student doesn't own
        - validate_all (validate_commits + validate_mrs): Also validates README location
        
        Args:
            students: List of student records to enrich
            gitlab_service: GitLabService instance
            target_date: Current date for calculating "this week"
            validate_commits: Whether to validate commit ownership
            validate_mrs: Whether to validate MRs and README ownership
            nofilter: Just check existence without ownership validation
            github_lookup: Dict mapping member_id to github username for ownership checks
        """
        from datetime import timezone
        
        current_date = target_date or datetime.now(timezone.utc)
        if not hasattr(current_date, 'tzinfo') or current_date.tzinfo is None:
            current_date = current_date.replace(tzinfo=timezone.utc)
        
        github_lookup = github_lookup or {}
        
        # Group students by readme_link to avoid duplicate API calls
        readme_to_students: Dict[str, List[StudentRecord]] = {}
        for student in students:
            readme_link = str(student.readme_link).strip() if student.readme_link else ""
            if readme_link:
                if readme_link not in readme_to_students:
                    readme_to_students[readme_link] = []
                readme_to_students[readme_link].append(student)
        
        enriched_count = 0
        ownership_flagged_count = 0
        readme_flagged_count = 0
        readme_nonexistent_count = 0
        readme_link_missing_count = 0
        mr_mismatch_count = 0
        
        # Determine mode string for logging
        if nofilter:
            mode_str = "nofilter (existence only)"
        elif validate_commits and validate_mrs:
            mode_str = "validate_all (ownership + README location)"
        elif validate_commits:
            mode_str = "validate_commits (ownership)"
        else:
            mode_str = "basic"
        
        print(f"[TrackerProcessor] Enriching {len(readme_to_students)} unique README links with GitLab data (mode: {mode_str})...")
        
        for readme_link, student_list in readme_to_students.items():
            # Get the mr_url and github username for ownership checks
            mr_url = ""
            student_github = ""
            student_member_id = ""
            for s in student_list:
                if s.mr_url and str(s.mr_url).strip():
                    mr_url = str(s.mr_url).strip()
                if s.member_id:
                    student_member_id = str(s.member_id).strip()
                    student_github = github_lookup.get(student_member_id, "").lower()
            
            # Fetch GitLab data with ownership info
            result = gitlab_service.enrich_student_data(
                readme_link=readme_link,
                mr_url=mr_url,
                owner_repo=None,
                current_date=current_date,
                validate_commits=True,  # Always fetch commit data
                validate_mrs=True,  # Always fetch MR data
                expected_owner=student_github if (validate_commits or validate_mrs) else None
            )
            
            if result.success:
                enriched_count += 1
                
                # Apply data to all students with this README link
                for student in student_list:
                    phase_num = self._get_phase_number(student.current_phase)
                    
                    # Only populate commit/MR fields if student is in Phase 3 or 4
                    if phase_num >= 3:
                        student.total_commits = result.commit_links_found
                        student.commits_this_week = result.commits_this_week
                        student.last_commit_date = result.last_commit_date
                        student.days_since_commit = result.days_since_commit
                    
                    # Only populate MR fields if student is in Phase 4
                    if phase_num >= 4:
                        student.mr_status = result.mr_status
                        student.mr_created_date = result.mr_created_date
                        student.comment_count = result.mr_comment_count
                        
                        # Check if MR URL matches what's in README (validation modes only)
                        if (validate_commits or validate_mrs) and not nofilter:
                            student_mr = str(student.mr_url).strip() if student.mr_url else ""
                            if student_mr and not result.mr_in_readme:
                                student.grade_status = "🔴 AT RISK"
                                if student.intervention_type:
                                    student.intervention_type += "\nMR_URL_MISMATCH"
                                else:
                                    student.intervention_type = "MR_URL_MISMATCH"
                                mr_mismatch_count += 1
                    
                    # Get this student's github username
                    s_member_id = str(student.member_id).strip() if student.member_id else ""
                    s_github = github_lookup.get(s_member_id, "").lower()
                    
                    # Ownership validation (validate_commits or validate_all)
                    if (validate_commits or validate_mrs) and s_github:
                        # Check if commits are on repos the student owns
                        if result.commits_not_owned > 0:
                            student.grade_status = "🔴 AT RISK"
                            if student.intervention_type:
                                student.intervention_type += "\nCOMMITS_NOT_OWNED"
                            else:
                                student.intervention_type = "COMMITS_NOT_OWNED"
                            ownership_flagged_count += 1
                    
                    # README location validation (validate_all only)
                    if validate_mrs and s_github:
                        # Check if README is on student's own repo
                        if not result.readme_owned_by_student:
                            student.grade_status = "🔴 AT RISK"
                            if student.intervention_type:
                                student.intervention_type += "\nREADME_NOT_OWNED"
                            else:
                                student.intervention_type = "README_NOT_OWNED"
                            readme_flagged_count += 1
            else:
                # README exists in link but file not found/accessible - flag as nonexistent
                print(f"[TrackerProcessor] Could not enrich {readme_link}: {result.error_message}")
                for student in student_list:
                    student.grade_status = "🔴 AT RISK"
                    if student.intervention_type:
                        student.intervention_type += "\nREADME_NONEXISTENT"
                    else:
                        student.intervention_type = "README_NONEXISTENT"
                    readme_nonexistent_count += 1
        
        # Flag students who have no readme_link at all
        for student in students:
            readme_link = str(student.readme_link).strip() if student.readme_link else ""
            if not readme_link:
                student.grade_status = "🔴 AT RISK"
                if student.intervention_type:
                    student.intervention_type += "\nREADME_LINK_MISSING"
                else:
                    student.intervention_type = "README_LINK_MISSING"
                readme_link_missing_count += 1
        
        print(f"[TrackerProcessor] GitLab enrichment complete: {enriched_count} READMEs processed")
        if readme_link_missing_count > 0:
            print(f"[TrackerProcessor] Flagged {readme_link_missing_count} students with README_LINK_MISSING (no link provided)")
        if readme_nonexistent_count > 0:
            print(f"[TrackerProcessor] Flagged {readme_nonexistent_count} students with README_NONEXISTENT (link invalid/inaccessible)")
        if mr_mismatch_count > 0:
            print(f"[TrackerProcessor] Flagged {mr_mismatch_count} students with MR_URL_MISMATCH")
        if ownership_flagged_count > 0:
            print(f"[TrackerProcessor] Flagged {ownership_flagged_count} students with COMMITS_NOT_OWNED")
        if readme_flagged_count > 0:
            print(f"[TrackerProcessor] Flagged {readme_flagged_count} students with README_NOT_OWNED")
    
    def _calculate_missed_deadlines(self, start_date: datetime, target_date: datetime) -> int:
        """Calculate the number of Wednesday and Sunday deadlines missed between dates.
        
        Args:
            start_date: Program start date
            target_date: Date to check up to
            
        Returns:
            Number of missed deadlines (Wednesdays + Sundays that have passed)
        """
        missed = 0
        
        # Get the weekday of start_date (Monday=0, Sunday=6)
        start_weekday = start_date.weekday()
        
        # Calculate first Wednesday and Sunday after start_date
        days_to_wed = (2 - start_weekday) % 7  # Wednesday = 2
        days_to_sun = (6 - start_weekday) % 7  # Sunday = 6
        if days_to_sun == 0:
            days_to_sun = 7  # If start is Sunday, next Sunday is 7 days
        
        first_wed = start_date + timedelta(days=days_to_wed)
        first_sun = start_date + timedelta(days=days_to_sun)
        
        # Count all Wednesdays that have passed
        current_wed = first_wed
        while current_wed.date() <= target_date.date():
            missed += 1
            current_wed += timedelta(days=7)
        
        # Count all Sundays that have passed
        current_sun = first_sun
        while current_sun.date() <= target_date.date():
            missed += 1
            current_sun += timedelta(days=7)
        
        return missed
    
    def _build_master_discord_lookup(self, master_data: bytes) -> Dict[str, str]:
        """Build a member_id -> discord_username lookup from master roster CSV.
        
        Args:
            master_data: Raw bytes of master CSV file
            
        Returns:
            Dict mapping member_id to discord_username
        """
        discord_lookup: Dict[str, str] = {}
        
        try:
            text_data = master_data.decode('utf-8-sig')
            text_data = _preprocess_master_csv(text_data)
            
            # Auto-detect delimiter (handles both CSV and TSV)
            sample = text_data[:4096]
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=',\t;|')
            except csv.Error:
                dialect = 'excel'
            
            csv_reader = csv.DictReader(io.StringIO(text_data), dialect=dialect)
            rows = list(csv_reader)
            
            if not rows:
                return discord_lookup
            
            headers = list(rows[0].keys())
            
            # Find member_id column
            member_id_col = self._find_column(headers, MASTER_CSV_COLUMNS["member_id"])
            if not member_id_col:
                print("[TrackerProcessor] Master CSV: Member ID column not found")
                return discord_lookup
            
            # Find discord username column
            discord_col = self._find_column(headers, MASTER_CSV_COLUMNS["discord_username"])
            if not discord_col:
                print("[TrackerProcessor] Master CSV: Discord Username column not found")
                return discord_lookup
            
            # Build lookup
            for row in rows:
                member_id = str(row.get(member_id_col, "")).strip()
                discord_username = str(row.get(discord_col, "")).strip()
                
                if member_id and discord_username:
                    discord_lookup[member_id] = discord_username
            
            print(f"[TrackerProcessor] Master CSV: Built lookup with {len(discord_lookup)} entries")
            
        except Exception as e:
            print(f"[TrackerProcessor] Error parsing master CSV: {e}")
        
        return discord_lookup
    
    def _build_master_contact_lookup(self, master_data: bytes) -> Dict[str, Dict[str, str]]:
        """Build a member_id -> contact info lookup from master roster CSV.
        
        Args:
            master_data: Raw bytes of master CSV file
            
        Returns:
            Dict mapping member_id to dict with discord, email, phone
        """
        contact_lookup: Dict[str, Dict[str, str]] = {}
        
        try:
            text_data = master_data.decode('utf-8-sig')
            text_data = _preprocess_master_csv(text_data)
            
            sample = text_data[:4096]
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=',\t;|')
            except csv.Error:
                dialect = 'excel'
            
            csv_reader = csv.DictReader(io.StringIO(text_data), dialect=dialect)
            rows = list(csv_reader)
            
            if not rows:
                return contact_lookup
            
            headers = list(rows[0].keys())
            
            member_id_col = self._find_column(headers, MASTER_CSV_COLUMNS["member_id"])
            if not member_id_col:
                return contact_lookup
            
            discord_col = self._find_column(headers, MASTER_CSV_COLUMNS["discord_username"])
            email_col = self._find_column(headers, MASTER_CSV_COLUMNS["email"])
            phone_col = self._find_column(headers, MASTER_CSV_COLUMNS["phone"])
            
            for row in rows:
                member_id = str(row.get(member_id_col, "")).strip()
                if member_id:
                    contact_lookup[member_id] = {
                        'discord': str(row.get(discord_col, "")).strip() if discord_col else "",
                        'email': str(row.get(email_col, "")).strip() if email_col else "",
                        'phone': str(row.get(phone_col, "")).strip() if phone_col else "",
                    }
            
            print(f"[TrackerProcessor] Master CSV: Built contact lookup with {len(contact_lookup)} entries")
            
        except Exception as e:
            print(f"[TrackerProcessor] Error parsing master CSV for contacts: {e}")
        
        return contact_lookup
    
    def _build_app_phone_lookup(self, app_data: bytes) -> Dict[str, str]:
        """Build a member_id -> phone lookup from app data CSV.
        
        Args:
            app_data: Raw bytes of app CSV file
            
        Returns:
            Dict mapping member_id to phone number
        """
        phone_lookup: Dict[str, str] = {}
        
        try:
            text_data = app_data.decode('utf-8-sig')
            text_data = _preprocess_master_csv(text_data)
            
            sample = text_data[:4096]
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=',\t;|')
            except csv.Error:
                dialect = 'excel'
            
            csv_reader = csv.DictReader(io.StringIO(text_data), dialect=dialect)
            rows = list(csv_reader)
            
            if not rows:
                return phone_lookup
            
            headers = list(rows[0].keys())
            
            member_id_col = self._find_column(headers, MASTER_CSV_COLUMNS["member_id"])
            if not member_id_col:
                print("[TrackerProcessor] App CSV: Member ID column not found")
                return phone_lookup
            
            phone_col = self._find_column(headers, MASTER_CSV_COLUMNS["phone"])
            if not phone_col:
                print("[TrackerProcessor] App CSV: Phone column not found")
                return phone_lookup
            
            for row in rows:
                member_id = str(row.get(member_id_col, "")).strip()
                phone = str(row.get(phone_col, "")).strip()
                
                if member_id and phone:
                    phone_lookup[member_id] = phone
            
            print(f"[TrackerProcessor] App CSV: Built phone lookup with {len(phone_lookup)} entries")
            
        except Exception as e:
            print(f"[TrackerProcessor] Error parsing app CSV for phones: {e}")
        
        return phone_lookup
    
    def _build_discord_lookup(self, raw_rows: List[Dict]) -> Dict[str, str]:
        """Build a member_id -> discord_username lookup from typeform data.
        
        Scans all rows to find the most recent discord username for each member_id.
        """
        discord_lookup: Dict[str, str] = {}
        
        if not raw_rows:
            return discord_lookup
        
        headers = list(raw_rows[0].keys()) if raw_rows else []
        
        # Find the discord username column
        discord_col = self._find_column(headers, DISCORD_USERNAME_COLUMNS)
        
        if not discord_col:
            return discord_lookup
        
        # Find member_id column
        member_id_col = None
        for col, field in CSV_COLUMN_MAP.items():
            if field == "member_id" and col in headers:
                member_id_col = col
                break
        
        if not member_id_col:
            return discord_lookup
        
        # Build lookup (later entries override earlier ones, giving most recent)
        for row in raw_rows:
            member_id = str(row.get(member_id_col, "")).strip()
            discord_username = str(row.get(discord_col, "")).strip()
            
            if member_id and discord_username:
                discord_lookup[member_id] = discord_username
        
        return discord_lookup
    
    def _build_name_lookup_from_master(self, master_data: bytes) -> Dict[str, str]:
        """Build a name -> member_id lookup from master CSV for fallback matching.
        
        Args:
            master_data: Raw bytes of master CSV file
            
        Returns:
            Dict mapping lowercase name to member_id
        """
        name_lookup: Dict[str, str] = {}
        
        try:
            text_data = master_data.decode('utf-8-sig')
            text_data = _preprocess_master_csv(text_data)
            
            sample = text_data[:4096]
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=',\t;|')
            except csv.Error:
                dialect = 'excel'
            
            csv_reader = csv.DictReader(io.StringIO(text_data), dialect=dialect)
            rows = list(csv_reader)
            
            if not rows:
                return name_lookup
            
            headers = list(rows[0].keys())
            
            # Find columns
            member_id_col = self._find_column(headers, MASTER_CSV_COLUMNS["member_id"])
            name_col = self._find_column(headers, MASTER_CSV_COLUMNS["full_name"])
            
            if not member_id_col or not name_col:
                return name_lookup
            
            # Build lookup
            for row in rows:
                member_id = str(row.get(member_id_col, "")).strip()
                name = str(row.get(name_col, "")).strip().lower()
                
                if member_id and name:
                    name_lookup[name] = member_id
            
        except Exception as e:
            print(f"[TrackerProcessor] Error building name lookup: {e}")
        
        return name_lookup
    
    def _transform_records(self, raw_rows: List[Dict], 
                          discord_lookup: Optional[Dict[str, str]] = None,
                          name_lookup: Optional[Dict[str, str]] = None,
                          contact_lookup: Optional[Dict[str, Dict[str, str]]] = None) -> List[StudentRecord]:
        """Transform raw CSV rows into StudentRecord objects.
        
        Args:
            raw_rows: List of raw CSV row dictionaries
            discord_lookup: Optional member_id -> discord_username mapping
            name_lookup: Optional name -> member_id mapping for fallback matching
            contact_lookup: Optional member_id -> {discord, email, phone} mapping
        """
        students = []
        discord_lookup = discord_lookup or {}
        contact_lookup = contact_lookup or {}
        
        for row in raw_rows:
            student = StudentRecord()
            student.raw_data = row
            
            for csv_col, field_name in CSV_COLUMN_MAP.items():
                value = _get_value_flexible(row, csv_col)
                if value is not None:
                    
                    # Handle special field mappings
                    if field_name == "_submission_type":
                        if "Wednesday" in value:
                            student.wed_submitted = True
                        elif "Sunday" in value:
                            student.sun_submitted = True
                    elif field_name == "_submission_day":
                        # Handle "Submission for" column with values like "Wed", "Sun"
                        val_lower = str(value).strip().lower()
                        if val_lower in ["wed", "wednesday"]:
                            student.wed_submitted = True
                        elif val_lower in ["sun", "sunday"]:
                            student.sun_submitted = True
                    elif field_name == "_tags":
                        # Check for AI Generated tag
                        if "AI Generated" in str(value):
                            student.cam_notes = "[AI Generated Response]"
                    elif field_name == "week":
                        # Extract week number from typeform input
                        try:
                            week_str = str(value).replace("Week ", "").strip()
                            student.week = int(week_str)
                        except:
                            student.week = 0
                    elif field_name == "contribution_num":
                        # Extract contribution number
                        try:
                            if "Contribution" in value:
                                num = value.split()[-1]
                                student.contribution_num = int(num)
                            elif "No Contribution" in value:
                                student.contribution_num = 0
                        except:
                            student.contribution_num = 1
                    elif field_name == "current_phase":
                        # Normalize phase names - only set if not already set AND value is not empty
                        # (handles duplicate columns where one may be empty)
                        if value and str(value).strip() and not student.current_phase:
                            student.current_phase = self._normalize_phase(value)
                    elif field_name in ["why_chosen_complete", "reproduction_complete", 
                                       "solution_complete", "implementation_complete",
                                       "testing_complete", "feedback_complete"]:
                        # Convert various true values to boolean
                        # Handle: 1, 1.0, "1", "1.0", Yes, TRUE, etc.
                        val_str = str(value).strip().lower()
                        is_true = val_str in ["1", "1.0", "yes", "true"]
                        setattr(student, field_name, is_true)
                    elif field_name == "blocked":
                        val_str = str(value).strip().lower()
                        setattr(student, field_name, val_str in ["1", "1.0", "yes", "true"])
                    elif hasattr(student, field_name):
                        setattr(student, field_name, value)
            
            # Check for invalid Member ID values (#N/A, empty, etc.)
            # And try fallback matching if primary Member ID is invalid
            member_id_col_value = _get_value_flexible(row, "Member ID")
            whats_member_id_value = _get_value_flexible(row, "What's your Member ID?")
            
            invalid_values = ['#N/A', 'N/A', 'NULL', 'NONE', '#REF!', '#VALUE!', '-', '']
            primary_is_invalid = (
                not member_id_col_value or 
                str(member_id_col_value).strip().upper() in invalid_values
            )
            secondary_is_invalid = (
                not whats_member_id_value or 
                str(whats_member_id_value).strip().upper() in invalid_values
            )
            
            if primary_is_invalid:
                # Primary Member ID is invalid - try fallbacks
                if not secondary_is_invalid:
                    # Use "What's your Member ID?" as fallback
                    student.member_id = str(whats_member_id_value).strip()
                    student.invalid_member_id = True
                elif name_lookup:
                    # Try name matching as last resort
                    name = None
                    for name_col in ["What's your name?", "Name", "name", "Full Name"]:
                        name = _get_value_flexible(row, name_col)
                        if name and str(name).strip():
                            break
                    
                    if name:
                        name_clean = str(name).strip().lower()
                        if name_clean in name_lookup:
                            student.member_id = name_lookup[name_clean]
                            student.invalid_member_id = True
                        else:
                            student.invalid_member_id = True
                    else:
                        student.invalid_member_id = True
                else:
                    student.invalid_member_id = True
            
            # Check for Member ID mismatch between columns (only if primary was valid)
            if not primary_is_invalid and member_id_col_value and whats_member_id_value:
                member_id_clean = str(member_id_col_value).strip()
                whats_clean = str(whats_member_id_value).strip()
                if member_id_clean and whats_clean and member_id_clean != whats_clean:
                    student.member_id_mismatch = True
            
            # Lookup contact info from master roster
            if student.member_id:
                member_id = str(student.member_id).strip()
                
                # Try contact lookup first (has all info)
                if member_id in contact_lookup:
                    contact = contact_lookup[member_id]
                    if not student.discord_username and contact.get('discord'):
                        student.discord_username = contact['discord']
                    if not student.email and contact.get('email'):
                        student.email = contact['email']
                    if not student.phone and contact.get('phone'):
                        student.phone = contact['phone']
                # Fall back to discord_lookup for backward compatibility
                elif member_id in discord_lookup and not student.discord_username:
                    student.discord_username = discord_lookup[member_id]
            
            # Calculate submission number: Wed W1=1, Sun W1=2, Wed W2=3, Sun W2=4, etc.
            if student.week > 0:
                if student.wed_submitted:
                    student.submission_num = (student.week - 1) * 2 + 1
                elif student.sun_submitted:
                    student.submission_num = (student.week - 1) * 2 + 2
            
            students.append(student)
        
        return students
    
    def _normalize_phase(self, phase_str: str) -> str:
        """Normalize phase string to consistent format (Phase # only)."""
        phase_str = str(phase_str).lower()
        
        if "1" in phase_str or "selection" in phase_str:
            return "Phase 1"
        elif "2" in phase_str or "reproduction" in phase_str:
            return "Phase 2"
        elif "3" in phase_str or "implementation" in phase_str:
            return "Phase 3"
        elif "4" in phase_str or "submission" in phase_str:
            return "Phase 4"
        return phase_str
    
    def _get_missing_deliverables(self, student: StudentRecord, phase_num: int) -> List[str]:
        """Get list of missing deliverables for a student's current phase.
        
        Args:
            student: The student record
            phase_num: The phase number (1-4)
            
        Returns:
            List of missing deliverable names
        """
        missing = []
        
        if phase_num == 1:
            if not (student.issue_url and str(student.issue_url).strip()):
                missing.append("issue_url")
            if not student.why_chosen_complete:
                missing.append("why_chosen_complete")
        elif phase_num == 2:
            if not (student.fork_url and str(student.fork_url).strip()):
                missing.append("fork_url")
            if not student.reproduction_complete:
                missing.append("reproduction_complete")
            if not student.solution_complete:
                missing.append("solution_complete")
        elif phase_num == 3:
            if not student.implementation_complete:
                missing.append("implementation_complete")
            if not student.testing_complete:
                missing.append("testing_complete")
        elif phase_num == 4:
            if not (student.mr_url and str(student.mr_url).strip()):
                missing.append("mr_url")
            if not student.feedback_complete:
                missing.append("feedback_complete")
        
        return missing
    
    def _apply_phase_completions(self, students: List[StudentRecord], 
                                 phase_completions: Dict[str, Dict]) -> None:
        """Apply manual phase completions to student records.
        
        If a student has a manual phase completion that's higher than their
        typeform-reported phase, update their current_phase to reflect the
        manual completion.
        
        Args:
            students: List of student records
            phase_completions: Dict mapping member_id to {phase: int, updated_at: str, updated_by: str}
        """
        if not phase_completions:
            return
        
        phase_names = {
            1: "Phase 1",
            2: "Phase 2",
            3: "Phase 3",
            4: "Phase 4"
        }
        
        for student in students:
            if not student.member_id:
                continue
            
            completion_data = phase_completions.get(student.member_id)
            if not completion_data:
                continue
            
            # Handle both list format (new) and int format (old)
            phases_data = completion_data.get('phases') or completion_data.get('phase')
            if isinstance(phases_data, list):
                manual_phase = max(phases_data) if phases_data else 0
            elif isinstance(phases_data, int):
                manual_phase = phases_data
            else:
                continue
            
            if manual_phase < 1 or manual_phase > 4:
                continue
            
            # Get current phase from typeform
            current_phase_num = self._get_phase_number(student.current_phase)
            
            # Only apply if manual phase is higher (more complete)
            if manual_phase > current_phase_num:
                student.current_phase = phase_names.get(manual_phase, f"Phase {manual_phase}")
                # Mark that this was manually set (optional - could add a flag field)
                if not hasattr(student, 'phase_manually_set'):
                    student.phase_manually_set = True
    
    def _calculate_derived_fields(self, students: List[StudentRecord], phase_completions: Dict = None) -> None:
        """Calculate derived fields for each student."""
        if phase_completions is None:
            phase_completions = {}
        # First, calculate weeks_in_phase by analyzing submission history
        self._calculate_weeks_in_phase(students, phase_completions)
        
        for student in students:
            # Calculate deliverables expected and complete based on current phase only
            phase_num = self._get_phase_number(student.current_phase)
            
            # Phase-specific deliverable requirements (not cumulative):
            # Phase 1 - 2 expected: issue_url, why_chosen_complete
            # Phase 2 - 3 expected: fork_url, reproduction_complete, solution_complete
            # Phase 3 - 2 expected: implementation_complete, testing_complete
            # Phase 4 - 2 expected: mr_url, feedback_complete
            
            if phase_num == 1:
                student.deliverables_expected = 2
                student.deliverables_complete = sum([
                    bool(student.issue_url and str(student.issue_url).strip()),
                    student.why_chosen_complete
                ])
            elif phase_num == 2:
                student.deliverables_expected = 3
                student.deliverables_complete = sum([
                    bool(student.fork_url and str(student.fork_url).strip()),
                    student.reproduction_complete,
                    student.solution_complete
                ])
            elif phase_num == 3:
                student.deliverables_expected = 2
                student.deliverables_complete = sum([
                    student.implementation_complete,
                    student.testing_complete
                ])
            elif phase_num == 4:
                student.deliverables_expected = 2
                student.deliverables_complete = sum([
                    bool(student.mr_url and str(student.mr_url).strip()),
                    student.feedback_complete
                ])
            else:
                student.deliverables_expected = 0
                student.deliverables_complete = 0
            
            # Calculate weeks remaining (assuming 10-week program)
            student.weeks_remaining = max(0, 10 - student.week)
            
            # Determine timeline type
            if student.weeks_remaining < 3 and phase_num < 3:
                student.timeline_type = "Compressed"
            elif student.weeks_remaining < 2 and phase_num < 4:
                student.timeline_type = "Critical"
            else:
                student.timeline_type = "Standard"
    
    def _calculate_weeks_in_phase(self, students: List[StudentRecord], phase_completions: Dict = None) -> None:
        """Calculate weeks_in_phase and submission_count_cumulative for each student.
        
        Groups submissions by member_id, sorts by week, and:
        - Counts consecutive weeks in the same phase
        - Counts cumulative complete submissions (all deliverables done)
        
        Args:
            students: List of StudentRecord objects
            phase_completions: Dict of member_id -> {phase: int, ...} for manual phase completions
        """
        if phase_completions is None:
            phase_completions = {}
        # Group submissions by member_id
        submissions_by_member: Dict[str, List[StudentRecord]] = {}
        for student in students:
            member_id = str(student.member_id).strip()
            if member_id:
                if member_id not in submissions_by_member:
                    submissions_by_member[member_id] = []
                submissions_by_member[member_id].append(student)
        
        # For each student, calculate weeks_in_phase and submission_count based on their history
        for member_id, member_submissions in submissions_by_member.items():
            # Sort by week (ascending)
            member_submissions.sort(key=lambda s: s.week)
            
            # Track phase history PER CONTRIBUTION
            # When contribution_num changes, reset phase tracking
            current_contribution = None
            current_phase = None
            phase_start_week = None
            contribution_start_week = None  # Track when each contribution started
            phases_submitted_for_contribution: set = set()  # Track which phases have submissions
            
            for submission in member_submissions:
                phase_num = self._get_phase_number(submission.current_phase)
                week = submission.week
                contrib_num = submission.contribution_num
                
                # Check if this is a new contribution
                if contrib_num != current_contribution:
                    # New contribution - reset phase tracking
                    current_contribution = contrib_num
                    current_phase = None
                    phase_start_week = None
                    contribution_start_week = week  # This contribution started this week
                    phases_submitted_for_contribution = set()  # Reset phases tracking
                
                # Track this phase as having a submission
                if phase_num > 0:
                    phases_submitted_for_contribution.add(phase_num)
                
                # Calculate weeks_in_phase
                if phase_num != current_phase:
                    # Phase changed (or first submission for this contribution)
                    previous_phase_num = current_phase
                    current_phase = phase_num
                    
                    if phase_start_week is None:
                        # First submission for this contribution
                        # Infer phase_start_week based on contribution_start_week
                        if phase_num == 1:
                            # Phase 1 - started when contribution started
                            phase_start_week = contribution_start_week
                        else:
                            # Higher phase on first submission of contribution
                            # Estimate based on weeks since contribution started
                            weeks_in_contribution = week - contribution_start_week + 1
                            # Assume ~1 week per earlier phase
                            estimated_phase_start = contribution_start_week + (phase_num - 1)
                            phase_start_week = min(week, max(contribution_start_week, estimated_phase_start))
                        submission.phase_changed_this_week = False
                        submission._unexpected_phase_change = False
                    else:
                        # We have history for this contribution - phase just changed
                        phase_start_week = week
                        submission.phase_changed_this_week = True
                        # Check for illogical phase change (going backwards)
                        submission._unexpected_phase_change = (phase_num < previous_phase_num)
                else:
                    # Same phase as previous submission
                    submission.phase_changed_this_week = False
                    submission._unexpected_phase_change = False
                
                # Calculate weeks in current phase
                weeks_in_current_phase = week - phase_start_week + 1
                submission.weeks_in_phase = max(1, weeks_in_current_phase)
                
                # Also track weeks on this contribution
                submission.weeks_on_contribution = week - contribution_start_week + 1
                submission.contribution_start_week = contribution_start_week
                
                # Check for missing IMMEDIATE previous phase (no submission record)
                # Only flag if the phase directly before current is missing
                # e.g., Phase 3 with Phase 2 missing = flag, but Phase 3 with Phase 1 missing (Phase 2 exists) = ok
                immediate_previous_phase = phase_num - 1
                
                # Check if the previous phase is covered by manual phase completion
                # phases can be stored as list (new format) or int (old format)
                manual_phases = []
                if member_id in phase_completions:
                    phases_data = phase_completions[member_id].get('phases') or phase_completions[member_id].get('phase')
                    if isinstance(phases_data, list):
                        manual_phases = phases_data
                    elif isinstance(phases_data, int):
                        # Old format: single int means all phases up to that number
                        manual_phases = list(range(1, phases_data + 1))
                
                # Combine all completed phases (from submissions + manual completions)
                all_completed_phases = phases_submitted_for_contribution | set(manual_phases)
                
                # MISSING_PREVIOUS_PHASE: immediate previous phase (current - 1) is missing
                previous_phase_covered = immediate_previous_phase in all_completed_phases
                
                if immediate_previous_phase >= 1 and not previous_phase_covered:
                    submission._missing_previous_phase = True
                    submission._missing_previous_phase_num = immediate_previous_phase
                else:
                    submission._missing_previous_phase = False
                    submission._missing_previous_phase_num = None
                
                # SKIPPED_PHASE: phases BEFORE the immediate previous that are missing
                # (but we have later phases completed - including current phase)
                # e.g., Phase 4 with phases [1, 4] = skipped 2, 3 missing is MISSING_PREVIOUS_PHASE
                # e.g., Phase 3 with phases [1, 3] = nothing skipped (missing 2 is MISSING_PREVIOUS_PHASE)
                # e.g., Phase 4 with phases [1, 3, 4] = skipped 2
                # e.g., Phase 4 with phases [4] = skipped 1, 2 (missing 3 is MISSING_PREVIOUS_PHASE)
                skipped_phases = []
                if phase_num >= 3:  # Only check for skipped if we're at phase 3+
                    # Check phases 1 to (current - 2) - these would be "skipped" not "missing previous"
                    for check_phase in range(1, immediate_previous_phase):
                        if check_phase not in all_completed_phases:
                            # Check if there's ANY later phase completed (including current phase)
                            # A phase is "skipped" if we've moved past it (current phase > check_phase + 1)
                            # and it wasn't completed
                            has_later_phase = any(p in all_completed_phases for p in range(check_phase + 1, phase_num + 1))
                            if has_later_phase:
                                skipped_phases.append(check_phase)
                
                submission._skipped_phases = skipped_phases
            
            # Calculate submission_count_cumulative
            # Count both Wednesday and Sunday submissions
            submission_count = 0
            for submission in member_submissions:
                if submission.wed_submitted:
                    submission_count += 1
                if submission.sun_submitted:
                    submission_count += 1
                submission.submission_count_cumulative = submission_count
            
            # Calculate consecutive_misses and track which submissions are missing
            # Build maps of week -> wed_submitted and week -> sun_submitted for this student
            week_to_wed_submitted: Dict[int, bool] = {}
            week_to_sun_submitted: Dict[int, bool] = {}
            for submission in member_submissions:
                week = submission.week
                # Track if any submission for this week has wed/sun submitted
                if week not in week_to_wed_submitted:
                    week_to_wed_submitted[week] = False
                if week not in week_to_sun_submitted:
                    week_to_sun_submitted[week] = False
                if submission.wed_submitted:
                    week_to_wed_submitted[week] = True
                if submission.sun_submitted:
                    week_to_sun_submitted[week] = True
            
            # Sort submissions: by week, then Wednesday before Sunday within same week
            def submission_sort_key(s):
                # Wednesday (wed_submitted=True) comes before Sunday (sun_submitted=True)
                # Lower number = earlier in sort
                sub_type = 0 if s.wed_submitted else 1
                return (s.week, sub_type)
            
            sorted_submissions = sorted(member_submissions, key=submission_sort_key)
            
            # Track if we've already "consumed" the consecutive misses for a gap
            # consecutive_misses should only be set on the FIRST entry after missing ones
            last_accounted_week = 0  # Week up to which misses have been accounted for
            
            for submission in sorted_submissions:
                current_week = submission.week
                consecutive_misses = 0
                
                # Check if previous week (current_week - 1) is missing Wed/Sun
                prev_week = current_week - 1
                if prev_week >= 1:
                    # Check if Wednesday was submitted for previous week
                    wed_submitted_prev = week_to_wed_submitted.get(prev_week, False)
                    # Check if Sunday was submitted for previous week
                    sun_submitted_prev = week_to_sun_submitted.get(prev_week, False)
                    
                    submission.last_week_wed_missing = not wed_submitted_prev
                    submission.last_week_sun_missing = not sun_submitted_prev
                
                # Only calculate consecutive_misses if we haven't already accounted for this gap
                if current_week > last_accounted_week:
                    # Count backwards from previous week
                    check_week = current_week - 1
                    while check_week >= 1:
                        if check_week in week_to_sun_submitted:
                            # We have data for this week
                            if not week_to_sun_submitted[check_week]:
                                # Missed this week's Sunday submission
                                consecutive_misses += 1
                            else:
                                # Found a Sunday submission - stop counting
                                break
                        else:
                            # No submission record for this week - count as a miss
                            consecutive_misses += 1
                        check_week -= 1
                    
                    # Mark that we've accounted for misses up to this week
                    if consecutive_misses > 0:
                        last_accounted_week = current_week
                
                submission.consecutive_misses = consecutive_misses
            
            # Track issue changes and contribution changes
            previous_issue_url = None
            previous_contribution = None
            previous_week = None
            issue_change_week = 0  # Track when issue was last changed
            
            for submission in member_submissions:
                week = submission.week
                current_issue = str(submission.issue_url).strip() if submission.issue_url else ""
                contrib_num = submission.contribution_num
                
                # issue_url_previous_week: Get issue URL from previous week's submission
                # Only set if there was a previous submission
                if previous_week is not None and previous_week == week - 1:
                    submission.issue_url_previous_week = previous_issue_url or ""
                else:
                    submission.issue_url_previous_week = ""
                
                # new_contribution_detected: Check if contribution number changed
                if previous_contribution is not None and contrib_num != previous_contribution:
                    submission.new_contribution_detected = True
                    # Reset issue tracking for new contribution
                    issue_change_week = 0
                    previous_issue_url = None
                else:
                    submission.new_contribution_detected = False
                
                # issue_changed: Check if issue URL changed from previous week
                if previous_issue_url and current_issue and current_issue != previous_issue_url:
                    submission.issue_changed = True
                    issue_change_week = week
                    
                    # issue_swap_detected: Same contribution but different issue
                    if contrib_num == previous_contribution:
                        submission.issue_swap_detected = True
                    else:
                        submission.issue_swap_detected = False
                else:
                    submission.issue_changed = False
                    submission.issue_swap_detected = False
                
                # issue_change_week: Set the week when issue was changed (persists)
                submission.issue_change_week = issue_change_week if issue_change_week > 0 else 0
                
                # Update tracking for next iteration
                if current_issue:  # Only update if there's an issue URL
                    previous_issue_url = current_issue
                previous_contribution = contrib_num
                previous_week = week
    
    def _get_phase_number(self, phase_str: str) -> int:
        """Extract phase number from phase string."""
        if "1" in phase_str:
            return 1
        elif "2" in phase_str:
            return 2
        elif "3" in phase_str:
            return 3
        elif "4" in phase_str:
            return 4
        return 0
    
    def _calculate_grade_status(self, students: List[StudentRecord], 
                                start_date: Optional[datetime] = None,
                                target_date: Optional[datetime] = None,
                                bypasses: Optional[Dict[str, Dict]] = None) -> None:
        """Calculate grade status and intervention type for each student.
        
        Args:
            students: List of student records to evaluate
            start_date: Program start date (for calculating per-week deadlines)
            target_date: Date the report is being run (for checking if deadlines passed)
            bypasses: Optional dict of bypassed submissions (key: "member_id:week")
        """
        bypasses = bypasses or {}
        
        # First pass: Build lookup of which students have Sunday submissions for each week
        # Key: (member_id, week) -> has_sunday
        student_week_has_sunday: Dict[Tuple[str, int], bool] = {}
        for s in students:
            key = (s.member_id or s.name, s.week)
            if s.sun_submitted:
                student_week_has_sunday[key] = True
            elif key not in student_week_has_sunday:
                student_week_has_sunday[key] = False
        
        # Second pass: Evaluate each student record
        for student in students:
            phase_num = self._get_phase_number(student.current_phase)
            
            # Check if this submission is bypassed
            bypass_key = f"{student.member_id}:{student.submission_num}"
            is_bypassed = student.submission_num > 0 and bypass_key in bypasses and bypasses[bypass_key].get('bypassed', False)
            
            # Note: We no longer automatically mark contribution_num > 1 as ON TRACK
            # because students may switch contributions without completing the previous one.
            # Instead, new_contribution_detected will be flagged for review.
            
            # Check if THIS student's week's Sunday deadline has passed
            sun_deadline_passed = False
            if start_date and target_date and student.week > 0:
                _, sun_deadline = self._get_week_deadlines(start_date, student.week)
                sun_deadline_passed = target_date.date() >= sun_deadline.date()
            
            # Check for AT RISK conditions
            at_risk = False
            intervention = ""
            
            # Check if this student has a Sunday submission for this week (across all rows)
            student_key = (student.member_id or student.name, student.week)
            student_has_sunday_for_week = student_week_has_sunday.get(student_key, False)
            
            # Missing both submissions (and no Sunday in another row for this week)
            # Don't flag if bypassed
            if not is_bypassed and not student.wed_submitted and not student.sun_submitted and not student_has_sunday_for_week:
                at_risk = True
                intervention = "MISSING_BOTH"
            
            # Has Wednesday but missing Sunday after Sunday deadline passed
            # Only flag if the student doesn't have a Sunday submission in another row
            # IMPORTANT: Still flag even if the Wednesday submission is bypassed - bypass only covers that specific submission
            elif student.wed_submitted and not student.sun_submitted and sun_deadline_passed and not student_has_sunday_for_week:
                at_risk = True
                intervention = "MISSING_SUNDAY"
            
            # SPECIAL CASE: Bypassed Wednesday submission but still missing Sunday
            # The bypass only covers the specific issue that was bypassed (e.g., STALLED), 
            # NOT the missing Sunday submission. Check explicitly.
            elif is_bypassed and not student.sun_submitted and sun_deadline_passed and not student_has_sunday_for_week:
                at_risk = True
                intervention = "MISSING_SUNDAY"
            
            # Phase critical (stuck in early phase late in program)
            elif student.week >= 6 and phase_num <= 2:
                at_risk = True
                intervention = "PHASE_COMPRESSED" if student.timeline_type == "Compressed" else "PHASE_CRITICAL"
            
            # Phase critical (less than 2 weeks remaining and not yet in Phase 4)
            elif student.weeks_remaining < 2 and phase_num < 4:
                at_risk = True
                intervention = "PHASE_COMPRESSED" if student.timeline_type == "Compressed" else "PHASE_CRITICAL"
            
            # Missed previous Sunday submission(s) - At Risk level
            # Wednesday is optional, only Sunday counts for consecutive misses
            elif student.consecutive_misses > 0 and student.last_week_sun_missing:
                at_risk = True
                intervention = "MISSING_SUNDAY"
            
            # Check for FLAGGED conditions
            flagged = False
            
            if not at_risk:
                # Stalled with blockers (moved from At Risk to Flagged)
                if student.blocked and student.blocker_desc:
                    flagged = True
                    intervention = "STALLED"
                
                # Check for missing immediate previous phase submission
                # Only flag if the phase directly before current has no submission record
                elif getattr(student, '_missing_previous_phase', False):
                    flagged = True
                    missing_phase_num = getattr(student, '_missing_previous_phase_num', None)
                    intervention = "MISSING_PREVIOUS_PHASE"
                    if missing_phase_num:
                        student._intervention_detail = f"Missing Phase {missing_phase_num}"
                
                # Check for skipped phases (earlier phases missing but later phases exist)
                # e.g., has phases 1, 3, 4 but missing phase 2 = skipped phase 2
                # Note: This can occur alongside MISSING_PREVIOUS_PHASE
                skipped_phases = getattr(student, '_skipped_phases', [])
                if skipped_phases:
                    flagged = True
                    # If already have missing previous phase, combine interventions
                    if intervention == "MISSING_PREVIOUS_PHASE":
                        skipped_str = ", ".join([str(p) for p in skipped_phases])
                        intervention = "MISSING_PREVIOUS_PHASE, SKIPPED_PHASE"
                        existing_detail = getattr(student, '_intervention_detail', '')
                        student._intervention_detail = f"{existing_detail}; Skipped Phase {skipped_str}"
                    else:
                        skipped_str = ", ".join([str(p) for p in skipped_phases])
                        intervention = "SKIPPED_PHASE"
                        student._intervention_detail = f"Skipped Phase {skipped_str}"
                
                # Check for illogical phase change (going backwards, e.g., Phase 3 -> Phase 2)
                elif getattr(student, '_unexpected_phase_change', False):
                    flagged = True
                    intervention = "UNEXPECTED_PHASE_CHANGE"
                
                # Check for new contribution started (switched contribution numbers)
                elif student.new_contribution_detected:
                    flagged = True
                    intervention = "NEW_CONTRIBUTION"
                    student._intervention_detail = f"Switched to Contribution {student.contribution_num}"
                
                # Check for issue change (student switched to a different issue)
                elif student.issue_changed:
                    flagged = True
                    intervention = "ISSUE_CHANGED"
                
                # Check for MR URL added in wrong phase (should only be in Phase 4)
                elif phase_num == 3 and student.mr_url and str(student.mr_url).strip():
                    flagged = True
                    intervention = "INCORRECT_PHASE_URL"
                
                # Missing deliverables for current phase (only check Sunday submissions)
                # Wednesday check-ins don't require complete deliverables
                elif student.sun_submitted and student.deliverables_complete < student.deliverables_expected:
                    flagged = True
                    missing_items = self._get_missing_deliverables(student, phase_num)
                    if missing_items:
                        items_list = "\n".join(f"-{item}" for item in missing_items)
                        intervention = f"MISSING_DELIVERABLES:\n{items_list}"
                    else:
                        intervention = "MISSING_DELIVERABLES"
                
                # No recent commits
                elif student.days_since_commit > 7:
                    flagged = True
                    intervention = "NO_ACTIVITY"
                
                # Compressed timeline
                elif student.timeline_type == "Compressed":
                    flagged = True
                    intervention = "TIMELINE_COMPRESSED"
            
            # Always append MEMBER_ID_MISMATCH if detected (data quality issue)
            # This is FLAGGED level (data quality concern, not critical)
            if student.member_id_mismatch:
                if intervention:
                    intervention += "\nMEMBER_ID_MISMATCH"
                else:
                    intervention = "MEMBER_ID_MISMATCH"
                # Member ID mismatch is FLAGGED level
                if not at_risk and not flagged:
                    flagged = True
            
            # Always append INVALID_MEMBER_ID if detected (data quality issue)
            # This is AT_RISK level since we can't verify the student's identity
            if student.invalid_member_id:
                if intervention:
                    intervention += "\nINVALID_MEMBER_ID"
                else:
                    intervention = "INVALID_MEMBER_ID"
                # Invalid Member ID is AT_RISK level
                if not at_risk:
                    at_risk = True
            
            # Set status
            if at_risk:
                student.grade_status = "🔴 AT RISK"
                # Note if the Wednesday submission was bypassed but Sunday is still missing
                if is_bypassed:
                    intervention = f"BYPASSED (Wed only)\n{intervention}"
            elif flagged:
                student.grade_status = "🟡 FLAGGED"
            elif is_bypassed:
                # Only mark as ON TRACK with BYPASSED if no other issues
                student.grade_status = "🟢 ON TRACK"
                intervention = "BYPASSED"
                student._bypassed = True
            else:
                student.grade_status = "🟢 ON TRACK"
            
            student.intervention_type = intervention
    
    def _create_master_tab(self, wb: Workbook, students: List[StudentRecord]) -> None:
        """Create Tab 1: Master Sheet with all student data."""
        ws = wb.create_sheet("Intervention Tracker")
        
        # Define all columns in order
        headers = [
            "member_id", "name", "discord_username", "email", "phone", "week",
            "submission_date", "wed_submitted", "sun_submitted", "submission_count_cumulative",
            "current_phase", "weeks_in_phase", "contribution_num", "contribution_start_week",
            "weeks_on_contribution", "weeks_remaining", "timeline_type", "phase_changed_this_week",
            "readme_link", "issue_url", "fork_url", "mr_url",
            "why_chosen_complete", "reproduction_complete", "solution_complete",
            "implementation_complete", "testing_complete", "feedback_complete",
            "deliverables_expected", "deliverables_complete",
            "commits_this_week", "last_commit_date", "days_since_commit", "total_commits",
            "mr_status", "mr_created_date", "comment_count", "has_maintainer_feedback",
            "progress_summary", "next_week_plan", "blocked", "blocker_desc", "support_requested",
            "issue_url_previous_week", "issue_changed", "issue_change_week",
            "issue_swap_detected", "new_contribution_detected",
            "grade_status", "intervention_type", "intervention_sent_date", "consecutive_misses",
            "tue_office_hours", "thu_office_hours", "wed_lecture", "cam_notes"
        ]
        
        # Write header row
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = Styles.HEADER_FILL
            cell.font = Styles.HEADER_FONT
            cell.alignment = Styles.CENTER_ALIGN
            cell.border = Styles.THIN_BORDER
        
        # Write data rows
        for row_idx, student in enumerate(students, 2):
            data = [
                student.member_id,
                student.name,
                student.discord_username,
                student.email,
                student.phone,
                student.week,
                student.submission_date,
                "Yes" if student.wed_submitted else "No",
                "Yes" if student.sun_submitted else "No",
                student.submission_count_cumulative,
                student.current_phase,
                student.weeks_in_phase,
                student.contribution_num,
                student.contribution_start_week,
                student.weeks_on_contribution,
                student.weeks_remaining,
                student.timeline_type,
                "Yes" if student.phase_changed_this_week else "No",
                student.readme_link,
                student.issue_url,
                student.fork_url,
                student.mr_url,
                "Yes" if student.why_chosen_complete else "No",
                "Yes" if student.reproduction_complete else "No",
                "Yes" if student.solution_complete else "No",
                "Yes" if student.implementation_complete else "No",
                "Yes" if student.testing_complete else "No",
                "Yes" if student.feedback_complete else "No",
                student.deliverables_expected,
                student.deliverables_complete,
                student.commits_this_week,
                student.last_commit_date,
                student.days_since_commit,
                student.total_commits,
                student.mr_status,
                student.mr_created_date,
                student.comment_count,
                "Yes" if student.has_maintainer_feedback else "No",
                student.progress_summary,
                student.next_week_plan,
                "Yes" if student.blocked else "No",
                student.blocker_desc,
                student.support_requested,
                student.issue_url_previous_week,
                "Yes" if student.issue_changed else "No",
                student.issue_change_week if student.issue_change_week else "",
                "Yes" if student.issue_swap_detected else "No",
                "Yes" if student.new_contribution_detected else "No",
                student.grade_status,
                student.intervention_type,
                student.intervention_sent_date,
                student.consecutive_misses,
                "✅" if student.tue_office_hours else "",
                "✅" if student.thu_office_hours else "",
                "✅" if student.wed_lecture else "",
                student.cam_notes
            ]
            
            # Determine row color based on grade status
            if student.grade_status == "🔴 AT RISK":
                row_fill = Styles.RED_FILL
            elif student.grade_status == "🟡 FLAGGED":
                row_fill = Styles.LIGHT_YELLOW_FILL
            else:
                row_fill = Styles.LIGHT_GREEN_FILL
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.fill = row_fill
                cell.border = Styles.THIN_BORDER
                cell.alignment = Styles.LEFT_ALIGN
        
        self._auto_fit_columns(ws)
        ws.freeze_panes = 'A2'
    
    def _get_student_priority_status(self, students: List[StudentRecord]) -> Tuple[Dict[str, str], Dict[str, str]]:
        """Determine each student's priority status across all their submissions.
        
        Logic: A student is forced to ON TRACK if any of these conditions are met:
        1. Has a Sunday submission that is ON TRACK
        2. Has contribution_num > 1 (on second or later contribution)
        3. Has an mr_url submitted (MR already created)
        
        Otherwise, use the worst status across all submissions.
        
        Args:
            students: List of all student records
            
        Returns:
            Tuple of:
            - Dict mapping member_id/name to their priority status
            - Dict mapping member_id/name to forced ON TRACK reason (empty string if not forced)
        """
        student_status: Dict[str, str] = {}
        student_forced_reason: Dict[str, str] = {}
        
        # Status priority (higher number = worse)
        status_priority = {
            "🟢 ON TRACK": 1,
            "🟡 FLAGGED": 2,
            "🔴 AT RISK": 3
        }
        
        # First pass: check if student meets any forced ON TRACK condition
        # Priority: Sunday ON TRACK > Has MR > Multiple Contributions
        # Note: BYPASSED is NOT a forced condition - it only applies to that specific submission,
        # other submissions should still be evaluated normally (worst status wins)
        for s in students:
            key = s.member_id or s.name
            
            # Skip if already has a higher-priority reason
            if key in student_forced_reason:
                continue
                
            # Condition 1: ON TRACK Sunday submission (highest priority - natural progression)
            if s.sun_submitted and s.grade_status == "🟢 ON TRACK":
                student_forced_reason[key] = ""  # Empty = natural ON TRACK, no special reason needed
            # Condition 2: Has submitted an MR (significant progress)
            elif s.mr_url and str(s.mr_url).strip():
                if key not in student_forced_reason:
                    student_forced_reason[key] = "Has MR submitted"
            # Condition 3: On second or later contribution (completed at least one)
            elif s.contribution_num > 1:
                if key not in student_forced_reason:
                    student_forced_reason[key] = f"On Contribution {s.contribution_num}"
        
        # Second pass: determine status
        for s in students:
            key = s.member_id or s.name
            
            # If student meets any forced ON TRACK condition, they're ON TRACK
            if key in student_forced_reason:
                student_status[key] = "🟢 ON TRACK"
                continue
            
            current_priority = status_priority.get(s.grade_status, 0)
            existing_priority = status_priority.get(student_status.get(key, ""), 0)
            
            # Update to worst (highest priority number) status
            if current_priority > existing_priority:
                student_status[key] = s.grade_status
        
        return student_status, student_forced_reason
    
    def _aggregate_student_issues(
        self, 
        students: List[StudentRecord], 
        status_filter: str,
        start_date: Optional[datetime] = None,
        target_date: Optional[datetime] = None
    ) -> List[Dict]:
        """Aggregate student records into unique entries with combined descriptions.
        
        Uses priority-based placement: a student only appears in the tab matching
        their WORST status across all submissions.
        
        Priority order: AT RISK > FLAGGED > ON TRACK
        
        Args:
            students: List of all student records
            status_filter: Grade status to filter by (e.g., "🔴 AT RISK")
            start_date: Program start date for deadline calculations
            target_date: Current target date to check against deadlines
            
        Returns:
            List of dicts with unique students and aggregated issue descriptions
        """
        # First, determine each student's priority status (worst across all submissions)
        priority_status, forced_reasons = self._get_student_priority_status(students)
        
        # Group by member_id, but only include students whose PRIORITY status matches the filter
        student_map: Dict[str, Dict] = {}
        
        for s in students:
            key = s.member_id or s.name
            
            # Skip if this student's priority status doesn't match the filter
            if priority_status.get(key) != status_filter:
                continue
            
            if key not in student_map:
                student_map[key] = {
                    'name': s.name,
                    'member_id': s.member_id,
                    'discord': s.discord_username,
                    'email': s.email,
                    'phone': s.phone,
                    'latest_week': s.week,
                    'latest_phase': s.current_phase,
                    'weeks_in_phase': s.weeks_in_phase,
                    'timeline': s.timeline_type,
                    'issues': set(),
                    'interventions': set(),
                    'submission_nums': set(),  # Track flagged submission numbers
                    'blocked': s.blocked,
                    'blocker_desc': s.blocker_desc,
                    'readme_link': s.readme_link,
                    'total_submissions': 0,
                    'deliverables': f"{s.deliverables_complete}/{s.deliverables_expected}",
                    'forced_on_track_reason': forced_reasons.get(key, ''),  # Why forced ON TRACK
                }
            
            # Update to latest week data, preferring Sunday submissions over Wednesday
            current_is_sunday = s.sun_submitted
            stored_week = student_map[key]['latest_week']
            should_update = (
                s.week > stored_week or  # Newer week always wins
                (s.week == stored_week and current_is_sunday)  # Same week: Sunday overrides Wednesday
            )
            
            if should_update:
                student_map[key]['latest_week'] = s.week
                student_map[key]['latest_phase'] = s.current_phase
                student_map[key]['weeks_in_phase'] = s.weeks_in_phase
                student_map[key]['timeline'] = s.timeline_type
                student_map[key]['blocked'] = s.blocked
                student_map[key]['blocker_desc'] = s.blocker_desc
                student_map[key]['deliverables'] = f"{s.deliverables_complete}/{s.deliverables_expected}"
            
            # Count submissions and track submission numbers
            if s.wed_submitted or s.sun_submitted:
                student_map[key]['total_submissions'] += 1
                if s.submission_num > 0:
                    student_map[key]['submission_nums'].add(s.submission_num)
            
            # Track which submissions exist per week (to avoid false "missing" reports)
            if 'week_submissions' not in student_map[key]:
                student_map[key]['week_submissions'] = {}
            if s.week not in student_map[key]['week_submissions']:
                student_map[key]['week_submissions'][s.week] = {'wed': False, 'sun': False}
            if s.wed_submitted:
                student_map[key]['week_submissions'][s.week]['wed'] = True
            if s.sun_submitted:
                student_map[key]['week_submissions'][s.week]['sun'] = True
            
            # Only collect issues for At Risk and Flagged students (not On Track)
            if status_filter == "🟢 ON TRACK":
                # For On Track students, don't collect issues - just count submissions
                continue
            
            # Collect all issues/interventions for At Risk and Flagged
            if s.intervention_type:
                student_map[key]['interventions'].add(s.intervention_type)
            
            # Add non-submission-related issues
            if s.intervention_type == "NO_SUBMISSIONS":
                if s.consecutive_misses > 0:
                    student_map[key]['issues'].add(f"No submissions ({s.consecutive_misses} deadlines missed)")
                else:
                    student_map[key]['issues'].add("No submissions")
            
            if s.consecutive_misses > 0:
                student_map[key]['issues'].add(f"Week {s.week}: {s.consecutive_misses} consecutive miss(es)")
            
            # Only flag missing deliverables for Sunday submissions
            if s.sun_submitted and s.deliverables_complete < s.deliverables_expected and s.week > 0:
                student_map[key]['issues'].add(f"Week {s.week}: Missing deliverables ({s.deliverables_complete}/{s.deliverables_expected})")
            
            if s.blocked:
                student_map[key]['issues'].add(f"Week {s.week}: Blocked - {s.blocker_desc[:50] if s.blocker_desc else 'Unknown'}")
            
            if s.timeline_type in ["Compressed", "Critical"] and s.week > 0:
                student_map[key]['issues'].add(f"Week {s.week}: {s.timeline_type} timeline")
            
            if getattr(s, '_unexpected_phase_change', False):
                student_map[key]['issues'].add(f"Week {s.week}: Unexpected phase change")
            
            if s.new_contribution_detected:
                student_map[key]['issues'].add(f"Week {s.week}: Switched to Contribution {s.contribution_num}")
            
            # Missing previous phase (immediate previous phase is missing)
            if getattr(s, '_missing_previous_phase', False):
                missing_num = getattr(s, '_missing_previous_phase_num', None)
                if missing_num:
                    student_map[key]['issues'].add(f"Week {s.week}: Missing Phase {missing_num}")
                else:
                    student_map[key]['issues'].add(f"Week {s.week}: Missing previous phase")
            
            # Skipped phases (earlier phases missing but later phases exist)
            skipped_phases = getattr(s, '_skipped_phases', [])
            if skipped_phases:
                skipped_str = ", ".join([str(p) for p in skipped_phases])
                student_map[key]['issues'].add(f"Week {s.week}: Skipped Phase {skipped_str}")
            
            if s.member_id_mismatch:
                student_map[key]['issues'].add("Member ID mismatch")
            
            if s.invalid_member_id:
                student_map[key]['issues'].add(f"Invalid Member ID: {s.member_id or 'empty'}")
        
        # Second pass: Add missing Sunday submission issues based on aggregated week data
        # Wednesday is optional (mid-week check-in), only Sunday is required
        # Only flag if the Sunday deadline has actually passed
        if status_filter != "🟢 ON TRACK":
            for key, data in student_map.items():
                week_subs = data.get('week_submissions', {})
                for week, subs in week_subs.items():
                    if week == 0:
                        continue  # Skip week 0 (no submissions students)
                    
                    # Only flag missing Sunday if we can verify the deadline has passed
                    if not subs['sun']:
                        should_flag = True
                        if start_date and target_date:
                            _, sun_deadline = self._get_week_deadlines(start_date, week)
                            # Only flag if Sunday deadline has passed
                            if target_date.date() < sun_deadline.date():
                                should_flag = False
                        
                        if should_flag:
                            data['issues'].add(f"Week {week}: Missing Sunday submission")
        
        # Convert to list and build descriptions
        result = []
        for key, data in student_map.items():
            # Build description from issues with chronological sorting
            # Sort by: week number first, then Wednesday before Sunday, then other issues
            def issue_sort_key(issue: str):
                # Extract week number if present
                week_match = re.search(r'Week (\d+)', issue)
                week_num = int(week_match.group(1)) if week_match else 0
                
                # Determine day order (Wednesday=1, Sunday=2, others=3)
                if 'Wednesday' in issue:
                    day_order = 1
                elif 'Sunday' in issue:
                    day_order = 2
                else:
                    day_order = 3
                
                return (week_num, day_order, issue)
            
            issues_list = sorted(data['issues'], key=issue_sort_key)
            # Use newline separator for interventions
            interventions = "\n".join(sorted(data['interventions'])) if data['interventions'] else "N/A"
            # Use newline separator for better readability
            description = "\n".join(issues_list) if issues_list else "No specific issues identified"
            # Format submission numbers as comma-separated sorted list
            submission_nums_str = ", ".join(str(n) for n in sorted(data['submission_nums'])) if data['submission_nums'] else ""
            
            result.append({
                **data,
                'interventions_str': interventions,
                'description': description,
                'submission_nums_str': submission_nums_str
            })
        
        return result
    
    def _create_at_risk_tab(
        self, 
        wb: Workbook, 
        students: List[StudentRecord],
        start_date: Optional[datetime] = None,
        target_date: Optional[datetime] = None
    ) -> None:
        """Create Tab 2: At Risk Students (unique per student with aggregated descriptions)."""
        ws = wb.create_sheet("P1 - At Risk")
        
        # Get unique students with aggregated issues
        at_risk = self._aggregate_student_issues(students, "🔴 AT RISK", start_date, target_date)
        
        # Sort: NO_SUBMISSIONS at the end, then by latest week (descending)
        def at_risk_sort_key(s):
            is_no_submission = "NO_SUBMISSION" in s.get('interventions_str', '')
            # Return tuple: (is_no_submission, -latest_week)
            # False (0) sorts before True (1), so non-NO_SUBMISSION comes first
            return (is_no_submission, -s['latest_week'])
        
        at_risk.sort(key=at_risk_sort_key)
        
        # Write header
        headers = ["Submission #", "Name", "Member ID", "Discord", "Email", "Phone", "Latest Week", "Phase", "Timeline",
                   "Deliverables", "Intervention Types", "Description"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = Styles.HEADER_FILL
            cell.font = Styles.HEADER_FONT
            cell.alignment = Styles.CENTER_ALIGN
            cell.border = Styles.THIN_BORDER
        
        # Write data
        for row_idx, student in enumerate(at_risk, 2):
            data = [
                student.get('submission_nums_str', ''),
                student['name'],
                student['member_id'],
                student['discord'],
                student.get('email', ''),
                student.get('phone', ''),
                student['latest_week'],
                student['latest_phase'],
                student['timeline'],
                student['deliverables'],
                student['interventions_str'],
                student['description']
            ]
            
            row_fill = Styles.RED_FILL
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.fill = row_fill
                cell.border = Styles.THIN_BORDER
                cell.alignment = Styles.LEFT_ALIGN
                if col == len(headers):  # Description column - wrap text
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        self._auto_fit_columns(ws)
        ws.freeze_panes = 'A2'
    
    def _create_flagged_tab(
        self, 
        wb: Workbook, 
        students: List[StudentRecord],
        start_date: Optional[datetime] = None,
        target_date: Optional[datetime] = None
    ) -> None:
        """Create Tab 3: Flagged Students (unique per student with aggregated descriptions)."""
        ws = wb.create_sheet("P2 - Flagged")
        
        # Get unique students with aggregated issues
        flagged = self._aggregate_student_issues(students, "🟡 FLAGGED", start_date, target_date)
        
        # Sort by intervention type (alphabetically), then by latest week (descending)
        flagged.sort(key=lambda s: (s.get('interventions_str', 'N/A'), -s['latest_week']))
        
        # Write header
        headers = ["Submission #", "Name", "Member ID", "Discord", "Email", "Phone", "Latest Week", "Phase", "Timeline",
                   "Deliverables", "Intervention Types", "Description"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = Styles.HEADER_FILL
            cell.font = Styles.HEADER_FONT
            cell.alignment = Styles.CENTER_ALIGN
            cell.border = Styles.THIN_BORDER
        
        # Write data
        for row_idx, student in enumerate(flagged, 2):
            data = [
                student.get('submission_nums_str', ''),
                student['name'],
                student['member_id'],
                student['discord'],
                student.get('email', ''),
                student.get('phone', ''),
                student['latest_week'],
                student['latest_phase'],
                student['timeline'],
                student['deliverables'],
                student['interventions_str'],
                student['description']
            ]
            
            row_fill = Styles.LIGHT_YELLOW_FILL
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.fill = row_fill
                cell.border = Styles.THIN_BORDER
                cell.alignment = Styles.LEFT_ALIGN
                if col == len(headers):  # Description column - wrap text
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        self._auto_fit_columns(ws)
        ws.freeze_panes = 'A2'
    
    def _create_on_track_tab(
        self, 
        wb: Workbook, 
        students: List[StudentRecord],
        start_date: Optional[datetime] = None,
        target_date: Optional[datetime] = None
    ) -> None:
        """Create Tab 4: On Track Students (unique per student with summary)."""
        ws = wb.create_sheet("P3 - On Track")
        
        # Get unique students with aggregated info
        on_track = self._aggregate_student_issues(students, "🟢 ON TRACK", start_date, target_date)
        
        # Helper to compute display description for sorting
        def get_display_description(s):
            forced_reason = s.get('forced_on_track_reason', '')
            if forced_reason:
                return f"✓ Forced ON TRACK: {forced_reason}"
            elif s['description'] != "No specific issues identified":
                return s['description']
            else:
                return "Progressing normally"
        
        # Sort by description (alphabetically), then by latest week (descending)
        on_track.sort(key=lambda s: (get_display_description(s), -s['latest_week']))
        
        # Write header
        headers = ["Submission #", "Name", "Member ID", "Discord", "Email", "Phone", "Latest Week", "Phase", 
                   "Total Submissions", "Deliverables", "Description"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = Styles.HEADER_FILL
            cell.font = Styles.HEADER_FONT
            cell.alignment = Styles.CENTER_ALIGN
            cell.border = Styles.THIN_BORDER
        
        # Write data
        for row_idx, student in enumerate(on_track, 2):
            # For on-track students, show positive status or forced reason
            description = get_display_description(student)
            
            data = [
                student.get('submission_nums_str', ''),
                student['name'],
                student['member_id'],
                student['discord'],
                student.get('email', ''),
                student.get('phone', ''),
                student['latest_week'],
                student['latest_phase'],
                student['total_submissions'],
                student['deliverables'],
                description
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.fill = Styles.LIGHT_GREEN_FILL
                cell.border = Styles.THIN_BORDER
                cell.alignment = Styles.LEFT_ALIGN
                if col == len(headers):  # Description column - wrap text
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        self._auto_fit_columns(ws)
        ws.freeze_panes = 'A2'
    
    def _create_summary_tab(self, wb: Workbook, students: List[StudentRecord]) -> None:
        """Create Tab 5: Weekly Summary Dashboard."""
        ws = wb.create_sheet("Weekly Summary")
        
        # Get unique students by member_id (use most recent week's record for each student)
        # Priority: Higher week > Sunday over Wednesday for same week (Sunday is more up-to-date)
        unique_students: Dict[str, StudentRecord] = {}
        for s in students:
            should_update = False
            if s.member_id not in unique_students:
                should_update = True
            else:
                current = unique_students[s.member_id]
                # Update if: newer week, OR same week but this is Sunday and current is not
                if s.week > current.week:
                    should_update = True
                elif s.week == current.week and s.sun_submitted and not current.sun_submitted:
                    should_update = True
            
            if should_update:
                unique_students[s.member_id] = s
        
        latest_records = list(unique_students.values())
        
        # Use the same priority logic as P1/P2/P3 tabs for consistency
        priority_status, _ = self._get_student_priority_status(students)
        
        # Calculate statistics based on priority status (matches P1/P2/P3 tabs)
        total = len(latest_records)
        on_track = sum(1 for key in priority_status.values() if key == "🟢 ON TRACK")
        flagged = sum(1 for key in priority_status.values() if key == "🟡 FLAGGED")
        at_risk = sum(1 for key in priority_status.values() if key == "🔴 AT RISK")
        
        # Count submissions from ALL rows (each submission row counts separately)
        sun_submitted = len([s for s in students if s.sun_submitted])
        wed_submitted = len([s for s in students if s.wed_submitted])
        
        phase_dist = {1: 0, 2: 0, 3: 0, 4: 0}
        for s in latest_records:
            phase_num = self._get_phase_number(s.current_phase)
            if phase_num in phase_dist:
                phase_dist[phase_num] += 1
        
        # Count unique students with MR URL (once per student)
        students_with_mr = set()
        students_with_merged_mr = set()
        for s in students:
            key = s.member_id or s.name
            if s.mr_url and str(s.mr_url).strip():
                students_with_mr.add(key)
            if s.mr_status and "merged" in s.mr_status.lower():
                students_with_merged_mr.add(key)
        mr_submitted = len(students_with_mr)
        mr_merged = len(students_with_merged_mr)
        
        # Interventions needed = At Risk + Flagged students (those who need attention)
        interventions_needed = at_risk + flagged
        
        # Get current week from data
        current_week = max(s.week for s in students) if students else 0
        
        # Create dashboard layout
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 20
        
        # Title
        ws.merge_cells('B2:C2')
        title_cell = ws.cell(row=2, column=2, value=f"WEEK {current_week} OVERVIEW")
        title_cell.fill = Styles.DASHBOARD_HEADER_FILL
        title_cell.font = Styles.DASHBOARD_TITLE_FONT
        title_cell.alignment = Styles.CENTER_ALIGN
        
        # Total students
        row = 4
        ws.cell(row=row, column=2, value="Total Students:").font = Styles.BOLD_FONT
        ws.cell(row=row, column=3, value=total)
        
        # Status breakdown
        row += 2
        ws.cell(row=row, column=2, value="🟢 On Track:").font = Styles.BOLD_FONT
        ws.cell(row=row, column=3, value=f"{on_track} ({on_track/total*100:.1f}%)" if total else "0")
        ws.cell(row=row, column=2).fill = Styles.GREEN_FILL
        
        row += 1
        ws.cell(row=row, column=2, value="🟡 Flagged:").font = Styles.BOLD_FONT
        ws.cell(row=row, column=3, value=f"{flagged} ({flagged/total*100:.1f}%)" if total else "0")
        ws.cell(row=row, column=2).fill = Styles.YELLOW_FILL
        
        row += 1
        ws.cell(row=row, column=2, value="🔴 At Risk:").font = Styles.BOLD_FONT
        ws.cell(row=row, column=3, value=f"{at_risk} ({at_risk/total*100:.1f}%)" if total else "0")
        ws.cell(row=row, column=2).fill = Styles.RED_FILL
        
        # Submissions section (counts all submission rows, not unique students)
        row += 2
        ws.merge_cells(f'B{row}:C{row}')
        section = ws.cell(row=row, column=2, value="Submissions (Total Rows)")
        section.fill = Styles.DASHBOARD_SECTION_FILL
        section.font = Styles.BOLD_FONT
        
        row += 1
        ws.cell(row=row, column=2, value="└─ Sunday:")
        ws.cell(row=row, column=3, value=f"{sun_submitted} submissions")
        
        row += 1
        ws.cell(row=row, column=2, value="└─ Wednesday:")
        ws.cell(row=row, column=3, value=f"{wed_submitted} submissions")
        
        # Phase distribution
        row += 2
        ws.merge_cells(f'B{row}:C{row}')
        section = ws.cell(row=row, column=2, value="Phase Distribution")
        section.fill = Styles.DASHBOARD_SECTION_FILL
        section.font = Styles.BOLD_FONT
        
        for phase in [1, 2, 3, 4]:
            row += 1
            ws.cell(row=row, column=2, value=f"└─ Phase {phase}:")
            ws.cell(row=row, column=3, value=f"{phase_dist[phase]} students")
        
        # MR section (unique students with MR URL)
        row += 2
        ws.cell(row=row, column=2, value="Students with MR:").font = Styles.BOLD_FONT
        ws.cell(row=row, column=3, value=f"{mr_submitted}/{total} ({mr_submitted/total*100:.1f}%)" if total else "0")
        
        row += 1
        ws.cell(row=row, column=2, value="MRs Merged:").font = Styles.BOLD_FONT
        ws.cell(row=row, column=3, value=f"{mr_merged} ({mr_merged/total*100:.1f}%)" if total else "0")
        
        # Interventions
        row += 2
        ws.cell(row=row, column=2, value="Interventions Needed:").font = Styles.BOLD_FONT
        ws.cell(row=row, column=3, value=interventions_needed)
        
        # Add border around dashboard
        for r in range(2, row + 1):
            for c in [2, 3]:
                cell = ws.cell(row=r, column=c)
                cell.border = Styles.THIN_BORDER
    
    def _auto_fit_columns(self, ws) -> None:
        """Auto-fit column widths."""
        for col_idx, column_cells in enumerate(ws.columns, 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            
            for cell in column_cells:
                try:
                    cell_length = len(str(cell.value or ""))
                    max_length = max(max_length, min(cell_length, 50))
                except:
                    pass
            
            ws.column_dimensions[column_letter].width = max_length + 2

