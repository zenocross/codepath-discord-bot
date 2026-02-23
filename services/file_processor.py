"""File processing services following SOLID principles.

This module provides extensible file processing capabilities:
- Abstract base class for processors (Open/Closed principle)
- CSV to Excel processor with styling
- File storage management
"""

import csv
import io
import os
from abc import ABC, abstractmethod
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, BinaryIO, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ==================== Data Classes ====================

@dataclass
class ProcessingResult:
    """Result of a file processing operation."""
    success: bool
    output_data: Optional[bytes] = None
    output_filename: Optional[str] = None
    error_message: Optional[str] = None
    rows_processed: int = 0


@dataclass
class StoredFile:
    """Metadata for a stored file."""
    filename: str
    filepath: Path
    uploaded_at: datetime
    user_id: int
    file_type: str


# ==================== Abstract Base Class (Interface Segregation) ====================

class FileProcessor(ABC):
    """Abstract base class for file processors.
    
    Implements the Open/Closed principle - open for extension,
    closed for modification. New processors can be added by
    subclassing without modifying existing code.
    """
    
    @property
    @abstractmethod
    def input_type(self) -> str:
        """Return the expected input file type (e.g., 'csv', 'xlsx')."""
        pass
    
    @property
    @abstractmethod
    def output_type(self) -> str:
        """Return the output file type (e.g., 'xlsx', 'csv')."""
        pass
    
    @abstractmethod
    def process(self, data: bytes, options: Optional[Dict[str, Any]] = None) -> ProcessingResult:
        """Process the input data and return the result.
        
        Args:
            data: Raw bytes of the input file
            options: Optional processing options
            
        Returns:
            ProcessingResult with output data or error
        """
        pass


# ==================== Concrete Processors ====================

class CsvToExcelProcessor(FileProcessor):
    """Converts CSV files to styled Excel files.
    
    Single Responsibility: Only handles CSV to Excel conversion with styling.
    """
    
    # Default colors (can be overridden via options)
    DEFAULT_HEADER_COLOR = "4472C4"  # Blue header
    DEFAULT_ROW_COLOR = "92D050"     # Green rows
    DEFAULT_ALT_ROW_COLOR = "C6EFCE"  # Light green alternating
    
    @property
    def input_type(self) -> str:
        return "csv"
    
    @property
    def output_type(self) -> str:
        return "xlsx"
    
    def process(self, data: bytes, options: Optional[Dict[str, Any]] = None) -> ProcessingResult:
        """Convert CSV to styled Excel.
        
        Options:
            row_color: Hex color for data rows (default: green)
            header_color: Hex color for header row (default: blue)
            alternating: Use alternating row colors (default: False)
            alt_row_color: Hex color for alternating rows
        """
        options = options or {}
        
        try:
            # Decode CSV data
            text_data = data.decode('utf-8-sig')  # Handle BOM if present
            csv_reader = csv.reader(io.StringIO(text_data))
            rows = list(csv_reader)
            
            if not rows:
                return ProcessingResult(
                    success=False,
                    error_message="CSV file is empty"
                )
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Processed Data"
            
            # Define styles
            header_fill = PatternFill(
                start_color=options.get('header_color', self.DEFAULT_HEADER_COLOR),
                end_color=options.get('header_color', self.DEFAULT_HEADER_COLOR),
                fill_type="solid"
            )
            header_font = Font(bold=True, color="FFFFFF")
            
            row_fill = PatternFill(
                start_color=options.get('row_color', self.DEFAULT_ROW_COLOR),
                end_color=options.get('row_color', self.DEFAULT_ROW_COLOR),
                fill_type="solid"
            )
            
            alt_row_fill = PatternFill(
                start_color=options.get('alt_row_color', self.DEFAULT_ALT_ROW_COLOR),
                end_color=options.get('alt_row_color', self.DEFAULT_ALT_ROW_COLOR),
                fill_type="solid"
            )
            
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            use_alternating = options.get('alternating', False)
            
            # Write data to worksheet
            for row_idx, row in enumerate(rows, start=1):
                for col_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                    
                    if row_idx == 1:
                        # Header row
                        cell.fill = header_fill
                        cell.font = header_font
                    else:
                        # Data rows
                        if use_alternating and row_idx % 2 == 0:
                            cell.fill = alt_row_fill
                        else:
                            cell.fill = row_fill
            
            # Auto-adjust column widths
            for col_idx, column_cells in enumerate(ws.columns, start=1):
                max_length = 0
                column_letter = get_column_letter(col_idx)
                
                for cell in column_cells:
                    try:
                        cell_length = len(str(cell.value or ""))
                        max_length = max(max_length, min(cell_length, 50))  # Cap at 50
                    except:
                        pass
                
                ws.column_dimensions[column_letter].width = max_length + 2
            
            # Freeze header row
            ws.freeze_panes = 'A2'
            
            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            return ProcessingResult(
                success=True,
                output_data=output.read(),
                output_filename="processed_data.xlsx",
                rows_processed=len(rows) - 1  # Exclude header
            )
            
        except UnicodeDecodeError as e:
            return ProcessingResult(
                success=False,
                error_message=f"Failed to decode CSV file: {e}"
            )
        except csv.Error as e:
            return ProcessingResult(
                success=False,
                error_message=f"Invalid CSV format: {e}"
            )
        except Exception as e:
            return ProcessingResult(
                success=False,
                error_message=f"Processing error: {e}"
            )


# ==================== File Storage Service (Single Responsibility) ====================

# Valid file category names
VALID_FILE_CATEGORIES = {"master", "typeform", "zoom"}


class FileStorageService:
    """Manages persistent file storage for processing.
    
    Single Responsibility: Only handles file storage and retrieval.
    Supports multiple named file categories (master, typeform, zoom).
    Persists file metadata to survive bot restarts.
    """
    
    def __init__(self, storage_dir: str = "data/uploads"):
        self.storage_dir = Path(storage_dir)
        self.storage_dir.mkdir(parents=True, exist_ok=True)
        self._files: Dict[str, Optional[StoredFile]] = {}
        self._load_all_metadata()
    
    def _get_metadata_file(self, category: str) -> Path:
        """Get metadata file path for a category."""
        return self.storage_dir / f"_{category}_file.json"
    
    def _load_all_metadata(self) -> None:
        """Load metadata for all file categories."""
        for category in VALID_FILE_CATEGORIES:
            self._files[category] = self._load_file_metadata(category)
        
        # Also load legacy "last_file" for backwards compatibility
        legacy_file = self._load_legacy_metadata()
        if legacy_file and not self._files.get("typeform"):
            self._files["typeform"] = legacy_file
    
    def _load_legacy_metadata(self) -> Optional[StoredFile]:
        """Load legacy _last_file.json for backwards compatibility."""
        legacy_path = self.storage_dir / "_last_file.json"
        if not legacy_path.exists():
            return None
        
        try:
            import json
            with open(legacy_path, 'r') as f:
                data = json.load(f)
            
            filepath = Path(data['filepath'])
            if not filepath.exists():
                return None
            
            return StoredFile(
                filename=data['filename'],
                filepath=filepath,
                uploaded_at=datetime.fromisoformat(data['uploaded_at']),
                user_id=data['user_id'],
                file_type=data['file_type']
            )
        except Exception as e:
            print(f"[FileStorage] Failed to load legacy metadata: {e}")
            return None
    
    def _load_file_metadata(self, category: str) -> Optional[StoredFile]:
        """Load file metadata for a specific category from disk."""
        metadata_file = self._get_metadata_file(category)
        if not metadata_file.exists():
            return None
        
        try:
            import json
            with open(metadata_file, 'r') as f:
                data = json.load(f)
            
            filepath = Path(data['filepath'])
            
            # Verify the file still exists
            if not filepath.exists():
                return None
            
            return StoredFile(
                filename=data['filename'],
                filepath=filepath,
                uploaded_at=datetime.fromisoformat(data['uploaded_at']),
                user_id=data['user_id'],
                file_type=data['file_type']
            )
        except Exception as e:
            print(f"[FileStorage] Failed to load {category} metadata: {e}")
            return None
    
    def _save_file_metadata(self, category: str, stored: StoredFile) -> None:
        """Save file metadata for a specific category to disk."""
        try:
            import json
            data = {
                'filename': stored.filename,
                'filepath': str(stored.filepath),
                'uploaded_at': stored.uploaded_at.isoformat(),
                'user_id': stored.user_id,
                'file_type': stored.file_type
            }
            with open(self._get_metadata_file(category), 'w') as f:
                json.dump(data, f, indent=2)
        except Exception as e:
            print(f"[FileStorage] Failed to save {category} metadata: {e}")
    
    def store_file(self, filename: str, data: bytes, user_id: int, 
                   category: Optional[str] = None) -> StoredFile:
        """Store a file and return its metadata.
        
        Args:
            filename: Original filename
            data: File data bytes
            user_id: Discord user ID who uploaded
            category: File category (master, typeform, zoom). If None, stores as generic.
        
        Note: If a file already exists for the category, it will be deleted first.
        """
        # Delete previous file for this category if it exists
        if category and category in VALID_FILE_CATEGORIES:
            previous = self._files.get(category)
            if previous and previous.filepath.exists():
                try:
                    previous.filepath.unlink()
                    print(f"[FileStorage] Deleted previous {category} file: {previous.filename}")
                except Exception as e:
                    print(f"[FileStorage] Error deleting previous {category} file: {e}")
        
        # Generate unique filename with timestamp and category prefix
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_filename = "".join(c for c in filename if c.isalnum() or c in "._-")
        category_prefix = f"{category}_" if category else ""
        stored_name = f"{category_prefix}{timestamp}_{safe_filename}"
        
        filepath = self.storage_dir / stored_name
        filepath.write_bytes(data)
        
        file_type = filename.rsplit('.', 1)[-1].lower() if '.' in filename else 'unknown'
        
        stored = StoredFile(
            filename=filename,
            filepath=filepath,
            uploaded_at=datetime.now(),
            user_id=user_id,
            file_type=file_type
        )
        
        if category and category in VALID_FILE_CATEGORIES:
            self._files[category] = stored
            self._save_file_metadata(category, stored)
        
        return stored
    
    def get_file(self, category: str) -> Optional[StoredFile]:
        """Get stored file for a specific category."""
        if category not in VALID_FILE_CATEGORIES:
            return None
        
        stored = self._files.get(category)
        # Verify the file still exists
        if stored and not stored.filepath.exists():
            self._files[category] = None
            return None
        return stored
    
    def get_last_file(self) -> Optional[StoredFile]:
        """Get the most recently stored typeform file (backwards compatibility)."""
        return self.get_file("typeform")
    
    def get_all_files(self) -> Dict[str, Optional[StoredFile]]:
        """Get all stored files by category."""
        return {
            category: self.get_file(category) 
            for category in VALID_FILE_CATEGORIES
        }
    
    def read_file(self, stored_file: StoredFile) -> bytes:
        """Read and return file contents."""
        return stored_file.filepath.read_bytes()
    
    def read_file_by_category(self, category: str) -> Optional[bytes]:
        """Read and return file contents for a specific category."""
        stored = self.get_file(category)
        if stored:
            return self.read_file(stored)
        return None
    
    def has_file(self, category: str) -> bool:
        """Check if a file exists for the given category."""
        return self.get_file(category) is not None
    
    def delete_file(self, category: str) -> bool:
        """Delete a stored file and its metadata for a specific category.
        
        Returns True if file was deleted, False if no file existed.
        """
        if category not in VALID_FILE_CATEGORIES:
            return False
        
        stored = self._files.get(category)
        if not stored:
            return False
        
        # Delete the actual file
        try:
            if stored.filepath.exists():
                stored.filepath.unlink()
        except Exception as e:
            print(f"[FileStorage] Error deleting {category} file: {e}")
        
        # Delete metadata file
        try:
            metadata_file = self._get_metadata_file(category)
            if metadata_file.exists():
                metadata_file.unlink()
        except Exception as e:
            print(f"[FileStorage] Error deleting {category} metadata: {e}")
        
        # Clear from memory
        self._files[category] = None
        return True
    
    def delete_all_files(self) -> int:
        """Delete all stored files and their metadata.
        
        Returns count of files deleted.
        """
        deleted = 0
        for category in VALID_FILE_CATEGORIES:
            if self.delete_file(category):
                deleted += 1
        return deleted
    
    def cleanup_old_files(self, max_age_hours: int = 24) -> int:
        """Remove files older than max_age_hours. Returns count of deleted files."""
        deleted = 0
        cutoff = datetime.now().timestamp() - (max_age_hours * 3600)
        
        for filepath in self.storage_dir.iterdir():
            # Skip metadata files
            if filepath.name.startswith("_") and filepath.name.endswith(".json"):
                continue
            if filepath.is_file() and filepath.stat().st_mtime < cutoff:
                filepath.unlink()
                deleted += 1
        
        return deleted
    
    # ==================== Tracker Settings Storage ====================
    
    def _get_settings_file(self) -> Path:
        """Get the tracker settings file path."""
        return self.storage_dir / "_tracker_settings.json"
    
    def _load_settings(self) -> Dict[str, Any]:
        """Load tracker settings from disk."""
        settings_file = self._get_settings_file()
        if not settings_file.exists():
            return {}
        
        try:
            import json
            with open(settings_file, 'r') as f:
                return json.load(f)
        except Exception as e:
            print(f"[FileStorage] Failed to load tracker settings: {e}")
            return {}
    
    def _save_settings(self, settings: Dict[str, Any]) -> None:
        """Save tracker settings to disk."""
        try:
            import json
            with open(self._get_settings_file(), 'w') as f:
                json.dump(settings, f, indent=2)
        except Exception as e:
            print(f"[FileStorage] Failed to save tracker settings: {e}")
    
    def set_start_date(self, date: datetime) -> None:
        """Set the program start date for week calculations."""
        settings = self._load_settings()
        settings['start_date'] = date.isoformat()
        self._save_settings(settings)
    
    def get_start_date(self) -> Optional[datetime]:
        """Get the program start date."""
        settings = self._load_settings()
        date_str = settings.get('start_date')
        if date_str:
            try:
                return datetime.fromisoformat(date_str)
            except:
                return None
        return None
    
    def set_last_submissions_date(self, date: datetime) -> None:
        """Set the last used submissions date for downloads."""
        settings = self._load_settings()
        settings['last_submissions_date'] = date.isoformat()
        self._save_settings(settings)
    
    def get_last_submissions_date(self) -> Optional[datetime]:
        """Get the last used submissions date."""
        settings = self._load_settings()
        date_str = settings.get('last_submissions_date')
        if date_str:
            try:
                return datetime.fromisoformat(date_str)
            except:
                return None
        return None


# ==================== Processor Registry (Dependency Inversion) ====================

class ProcessorRegistry:
    """Registry for file processors.
    
    Implements Dependency Inversion - high-level modules depend on
    this abstraction rather than concrete processors.
    """
    
    def __init__(self):
        self._processors: Dict[str, FileProcessor] = {}
    
    def register(self, name: str, processor: FileProcessor) -> None:
        """Register a processor by name."""
        self._processors[name] = processor
    
    def get(self, name: str) -> Optional[FileProcessor]:
        """Get a processor by name."""
        return self._processors.get(name)
    
    def get_by_input_type(self, input_type: str) -> Optional[FileProcessor]:
        """Get a processor that handles the given input type."""
        for processor in self._processors.values():
            if processor.input_type == input_type:
                return processor
        return None
    
    def list_processors(self) -> List[str]:
        """List all registered processor names."""
        return list(self._processors.keys())


# ==================== Default Registry Setup ====================

def create_default_registry() -> ProcessorRegistry:
    """Create a registry with default processors."""
    registry = ProcessorRegistry()
    registry.register('csv_to_excel', CsvToExcelProcessor())
    return registry

